from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import style  # 引入樣式模組

def create_work_hours_sheet(wb):
    sheet_name = "類別一-工作時數(員工)"
    ws = wb.create_sheet(title=sheet_name)
    
    # 設定標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    merged_cell = ws.cell(row=1, column=1, value="員工平均工作時數")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill  # 黃色背景
    
    # 第二行：標題列
    headers = ['月份', '員工數', '每日工時', '每月工作日數', '總加班時數', '總病假時數', '總事假時數', '總出差時數', '總婚喪時數', '總特休時數', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 灰色背景
    
    # 填充月份與基本數據
    months = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月']
    data = [
        {'employee_number': 23, 'daily_hours': 8, 'workday': 21},
        {'employee_number': 23, 'daily_hours': 8, 'workday': 15},
        {'employee_number': 22, 'daily_hours': 8, 'workday': 23},
        {'employee_number': 22, 'daily_hours': 8, 'workday': 19},
        {'employee_number': 24, 'daily_hours': 8, 'workday': 21},
        {'employee_number': 23, 'daily_hours': 8, 'workday': 21},
        {'employee_number': 24, 'daily_hours': 8, 'workday': 20},
        {'employee_number': 24, 'daily_hours': 8, 'workday': 23},
        {'employee_number': 23, 'daily_hours': 8, 'workday': 21},
        {'employee_number': 23, 'daily_hours': 8, 'workday': 20},
        {'employee_number': 25, 'daily_hours': 8, 'workday': 22},
        {'employee_number': 25, 'daily_hours': 8, 'workday': 22},
    ]
    
    for i, (month, info) in enumerate(zip(months, data), start=3):
        ws.cell(row=i, column=1, value=month)  # A 欄：月份
        ws.cell(row=i, column=2, value=info['employee_number'])  # B 欄：員工數
        ws.cell(row=i, column=3, value=info['daily_hours'])  # C 欄：每日工時
        ws.cell(row=i, column=4, value=info['workday'])  # D 欄：每月工作日數
    
    # 設定總計行
    total_row = len(months) + 3
    ws.cell(row=total_row, column=1, value="總計")
    ws.cell(row=total_row, column=2, value="=SUM(B3:B14)")
    ws.cell(row=total_row, column=4, value="=SUM(D3:D14)")
    
    # 套用框線
    for row in ws.iter_rows(min_row=2, max_row=total_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border
    
    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2
