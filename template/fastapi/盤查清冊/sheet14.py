from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import style # 引入樣式模組

def create_power_usage_sheet(wb):
    """建立 '類別二-電力使用量' 工作表"""
    sheet_name = "類別二-電力使用量"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=1, value="電力使用量")
    title_cell.alignment = style.center_alignment
    title_cell.fill = style.yellow_fill  # 標題背景色

    # 第二行：表頭
    headers = ['收據月份', '用電期間(起)', '用電期間(迄)', '填寫類型', '尖峰度數', '半尖峰度數', '周六半尖峰度數', '離峰度數', '當月總用電量或總金額', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 灰色背景

    # 預設數據
    data = [
        {'month': '2月', 'start_date': '請輸入西元/月/日', 'end_date': '請輸入西元/月/日', 'type': '', 'peak': 0, 'semi_peak': 0, 'sat_semi_peak': 0, 'off_peak': 0, 'total': 0, 'remark': 'none'},
        {'month': '', 'start_date': '請輸入西元/月/日', 'end_date': '請輸入西元/月/日', 'type': '', 'peak': 0, 'semi_peak': 0, 'sat_semi_peak': 0, 'off_peak': 0, 'total': 0, 'remark': 'none'},
        {'month': '', 'start_date': '請輸入西元/月/日', 'end_date': '請輸入西元/月/日', 'type': '', 'peak': 0, 'semi_peak': 0, 'sat_semi_peak': 0, 'off_peak': 0, 'total': 0, 'remark': 'none'}
    ]

    row = 3  # 從第三行開始寫入資料
    for item in data:
        for col_idx, key in enumerate(['month', 'start_date', 'end_date', 'type', 'peak', 'semi_peak', 'sat_semi_peak', 'off_peak', 'total', 'remark'], start=1):
            ws.cell(row=row, column=col_idx, value=item.get(key, ''))
        row += 1

    # 第 6 行：欄位說明
    notes = ['', '日期', '日期', '選擇', '數字', '數字', '數字', '數字', '數字', '可不填']
    for col_idx, note in enumerate(notes, start=1):
        ws.cell(row=row, column=col_idx, value=note)

    # 套用邊框
    for row_cells in ws.iter_rows(min_row=2, max_row=len(data)+2, min_col=1, max_col=len(headers)):
        for cell in row_cells:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2