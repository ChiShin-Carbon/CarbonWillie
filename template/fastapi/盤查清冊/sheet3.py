from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import style  # 引入樣式模組

def create_fire_extinguisher_sheet(wb):
    """建立 '類別一-滅火器' 工作表"""
    sheet_name = "類別一-滅火器"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    merged_cell = ws.cell(row=1, column=1, value="滅火器")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill  # 標題背景色

    # 合併第二列說明文字
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=2, column=1, value="公司本年度二氧化碳滅火器之請購及換藥紀錄（無則免填）").alignment = style.center_alignment

    # 定義表頭
    headers = ['品名', '成分', '規格(重量)', '單位', '使用量(支)', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 標題背景色

    # 填入資料
    data = [
        {'item_name': '乾粉滅火器', 'ingredient': '其他', 'specification': '6.5', 'unit': '公斤', 'usage': '148', 'remark': '5樓'},
        {'item_name': '二氧化碳滅火器', 'ingredient': 'CO2', 'specification': '4.5', 'unit': '公斤', 'usage': '11', 'remark': '5樓'},
        {'item_name': '乾粉滅火器', 'ingredient': '其他', 'specification': '6.5', 'unit': '公斤', 'usage': '77', 'remark': '5樓'},
        {'item_name': '二氧化碳滅火器', 'ingredient': 'CO2', 'specification': '4.5', 'unit': '公斤', 'usage': '1', 'remark': '5樓'},
        {'item_name': '乾粉滅火器', 'ingredient': '其他', 'specification': '3.5', 'unit': '公斤', 'usage': '257', 'remark': '5樓'},
        {'item_name': 'FM-200', 'ingredient': 'HFC-227ea', 'specification': '4.5', 'unit': '公斤', 'usage': '60', 'remark': '5樓'}
    ]
    
    row = 4
    for item in data:
        for col_idx, key in enumerate(['item_name', 'ingredient', 'specification', 'unit', 'usage', 'remark'], start=1):
            ws.cell(row=row, column=col_idx, value=item.get(key, ''))
        row += 1

    # 欄位備註
    note_row = row
    notes = ['文字', '選擇', '數字', '選擇', '數字', '文字']
    for col_idx, note in enumerate(notes, start=1):
        ws.cell(row=note_row, column=col_idx, value=note)

    # 套用黑色邊框
    for row in ws.iter_rows(min_row=3, max_row=len(data)+3, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        adjusted_width = (max_length + 4) * 1.2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
