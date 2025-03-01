from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 定義框線樣式（黑色、細線）
black_border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


# 建立 Excel 檔案
wz = Workbook()
#分工說明表
wa = wz.active # 選擇第一個工作表
wa.title = "分工說明表"
# 資料
data = [
    {
      #   'sheet': '類別一-公務車(汽油)',
        'department':'業務部',
        'name': '陳欣柔',
        'email': 'chen@gmail.com',
        'phone': '09722222',
        'address': 'admin',
        'password': 'chen',
        'role': 'admin',
        'remark': 'none'
     },
     {
      #   'sheet': '類別一-公務車(柴油)',
        'department':'',
        'name': '',
        'email': '',
        'phone': '',
        'address': '',
        'password': '',
        'role': '',
        'remark': '' 
     },  
     {
      #   'sheet': '類別一-滅火器',
        'department':'',
        'name': '',
        'email': '',
        'phone': '',
        'address': '',
        'password': '',
        'role': '',
        'remark': '' 
     },    
     {
      #   'sheet': '類別一-工作時數(員工)',
        'department':'',
        'name': '',
        'email': '',
        'phone': '',
        'address': '',
        'password': '',
        'role': '',
        'remark': '' 
     },
    {
      #   'sheet': '類別一-工作時數(非員工)',
        'department':'',
        'name': '',
        'email': '',
        'phone': '',
        'address': '',
        'password': '',
        'role': '',
        'remark': '' 
     },
    {
      #   'sheet': '類別一-冷媒',
        'department':'',
        'name': '',
        'email': '',
        'phone': '',
        'address': '',
        'password': '',
        'role': '',
        'remark': '' 
     },
    {
      #   'sheet': '類別一-固定式',
        'department':'',
        'name': '',
        'email': '',
        'phone': '',
        'address': '',
        'password': '',
        'role': '',
        'remark': '' 
     } ,
    {
      #   'sheet': '類別一-產生溫室氣體的排放製程',
        'department':'',
        'name': '',
        'email': '',
        'phone': '',
        'address': '',
        'password': '',
        'role': '',
        'remark': '' 
     } ,
    {
      #   'sheet': '類別一-廠內機具',
        'department':'',
        'name': '',
        'email': '',
        'phone': '',
        'address': '',
        'password': '',
        'role': '',
        'remark': '' 
     } ,
    {
      #   'sheet': '類別一-緊急發電機',
        'department':'',
        'name': '',
        'email': '',
        'phone': '',
        'address': '',
        'password': '',
        'role': '',
        'remark': '' 
     } 
     ]
# 定義標題
headers1 = ['表單名稱', '部門', '姓名(主要填寫人)', '電子郵件', '電話及分機號碼', '平台帳號', '平台密碼','平台權限', '備註']
for col_idx, header in enumerate(headers1, start=1):
    cell = wa.cell(row=1, column=col_idx, value=header)
    cell.alignment = Alignment(horizontal="center", vertical="center") # 置中
    cell.fill = PatternFill(fill_type="solid", fgColor="ffffcc00") # 設定儲存格的背景樣式
# 個別負責表單
sheet = ['類別一-公務車(汽油)','類別一-公務車(柴油)','類別一-滅火器','類別一-工作時數(員工)','類別一-工作時數(非員工)','類別一-冷媒','類別一-固定式','類別一-產生溫室氣體的排放製程','類別一-廠內機具','類別一-緊急發電機','類別一-焊條','類別一-氣體斷路器(GCB)','類別一-其他','類別二-電力使用量','類別二-間接蒸氣(汽電共生廠有做溫室氣體盤查)','類別二-間接蒸氣(汽電共生廠沒有做溫室氣體盤查)','類別二-其他']
for i, item in enumerate(sheet, start=2):  # i 對應 Excel 的行號（從 1 開始）
    wa.cell(row=i, column=1, value=item)  # 第一列（A欄）填入資料
# 資料開始寫入
row = 2  # 從第二行開始寫入資料
for person in data:
    for col_idx, value in enumerate(person.values(), start=2):  # 從 column=2 開始
        wa.cell(row=row, column=col_idx, value=value)
    row += 1  # 移動到下一行
# 套用黑框到 A1 到 I18
for row in wa.iter_rows(min_row=1, max_row=len(sheet)+1, min_col=1, max_col=len(headers1)):
    for cell in row:
        cell.border = black_border
# 自動調整欄寬
for col_idx, col_cells in enumerate(wa.columns, start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)  # 取得欄位字母 (如 A, B, C...)
    for cell in col_cells:
        if cell.value:
            # 取得最大字元長度
            max_length = max(max_length, len(str(cell.value)))
    # 預估調整寬度（+2 以增加間距）
    adjusted_width = (max_length + 4) * 1.5  # 乘以 1.2 來更準確反映 Excel 的寬度
    # 針對「表單名稱」欄 (A 欄) 再加長**
    if col_idx == 1:
        adjusted_width += 12
    wa.column_dimensions[col_letter].width = adjusted_width  # 設定欄位寬度


# 類別一-公務車(汽油)
wz.create_sheet(sheet[0])  # 插入工作表2 在最後方
wb = wz[sheet[0]] 
wb.title = sheet[0]
# 資料
data2 = [
    {
      #   'Doc_date': '1月',
        # 'oil_species': '車用汽油',
        'unit': '公升',
        'money': '1,000',
        'liters': '1,500',
        'remark': 'none', 
     },
     {
      #   'Doc_date': '2月',
        # 'oil_species': '車用汽油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     },  
     {
      #   'Doc_date': '3月',
        # 'oil_species': '車用汽油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     },    
     {
      #   'Doc_date': '4月',
        # 'oil_species': '車用汽油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     },
    {
      #   'Doc_date': '5月',
        # 'oil_species': '車用汽油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     },
    {
      #   'Doc_date': '6月',
        # 'oil_species': '車用汽油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     },
    {
      #   'Doc_date': '7月',
        # 'oil_species': '車用汽油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     } ,
    {
      #   'Doc_date': '8月',
        # 'oil_species': '車用汽油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     } ,
    {
      #   'Doc_date': '9月',
        # 'oil_species': '車用汽油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     } ,
    {
      #   'Doc_date': '10月',
        # 'oil_species': '車用汽油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     } 
     ]
# 定義標題
# 合併 A1 到 F1 儲存格
wb.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
merged_cell_1 = wb.cell(row=1, column=1, value=sheet[0])  # 設定合併後的內容
merged_cell_1.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
merged_cell_1.fill = PatternFill(fill_type="solid", fgColor="ffffcc00") # 設定儲存格的背景樣式
# 第二行
wb.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
# 第三行
headers2 = ['發票日期', '油種', '單位', '金額', '公升數', '備註']
for col_idx, header in enumerate(headers2, start=1):
    cell2 = wb.cell(row=3, column=col_idx, value=header)
    cell2.fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD") # 設定儲存格的背景樣式
# 月份
Doc_date = ['1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']
for i, item in enumerate(Doc_date, start=4):  # i 對應 Excel 的行號（從 1 開始）
    wb.cell(row=i, column=1, value=item)  # 第一列（A欄）填入資料
    wb.cell(row=i, column=2, value='車用汽油')  # 第二列（B欄）填入資料
# 資料開始寫入
row = 4  # 從第四行開始寫入資料
for information in data2:
    for col_idx, value in enumerate(information.values(), start=3):  # 從 column=3 開始
        wb.cell(row=row, column=col_idx, value=value)
    row += 1  # 移動到下一行
# 套用黑框到 A3 到 F15
for row in wb.iter_rows(min_row=3, max_row=len(Doc_date)+3, min_col=1, max_col=len(headers2)):
    for cell in row:
        cell.border = black_border
# 自動調整欄寬
for col_idx, col_cells in enumerate(wb.columns, start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)  # 取得欄位字母 (如 A, B, C...)
    for cell in col_cells:
        if cell.value:
            # 取得最大字元長度
            max_length = max(max_length, len(str(cell.value)))
    # 預估調整寬度（+2 以增加間距）
    adjusted_width = (max_length + 4) * 1.2  # 乘以 1.2 來更準確反映 Excel 的寬度
    wb.column_dimensions[col_letter].width = adjusted_width  # 設定欄位寬度


# 類別二-公務車(柴油)
wz.create_sheet(sheet[1])  # 插入工作表3 在最後方
wc = wz[sheet[1]] 
wc.title = sheet[1]
# 資料
data3 = [
    {
      #   'Doc_date': '1月',
        # 'oil_species': '車用柴油',
        'unit': '公升',
        'money': '1,000',
        'liters': '1,500',
        'remark': 'none', 
     },
     {
      #   'Doc_date': '2月',
        # 'oil_species': '車用柴油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     },  
     {
      #   'Doc_date': '3月',
        # 'oil_species': '車用柴油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     },    
     {
      #   'Doc_date': '4月',
        # 'oil_species': '車用柴油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     },
    {
      #   'Doc_date': '5月',
        # 'oil_species': '車用柴油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     },
    {
      #   'Doc_date': '6月',
        # 'oil_species': '車用柴油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     },
    {
      #   'Doc_date': '7月',
        # 'oil_species': '車用柴油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     } ,
    {
      #   'Doc_date': '8月',
        # 'oil_species': '車用柴油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     } ,
    {
      #   'Doc_date': '9月',
        # 'oil_species': '車用柴油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     } ,
    {
      #   'Doc_date': '10月',
        # 'oil_species': '車用柴油',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     } 
     ]
# 定義標題
# 合併 A1 到 F1 儲存格
wc.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
merged_cell_2 = wc.cell(row=1, column=1, value=sheet[1])  # 設定合併後的內容
merged_cell_2.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
merged_cell_2.fill = PatternFill(fill_type="solid", fgColor="ffffcc00") # 設定儲存格的背景樣式
# 第二行
wc.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
# 第三行
headers2 = ['發票日期', '油種', '單位', '金額', '公升數', '備註']
for col_idx, header in enumerate(headers2, start=1):
    cell2 = wc.cell(row=3, column=col_idx, value=header)
    cell2.fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD") # 設定儲存格的背景樣式
# 月份
Doc_date = ['1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']
for i, item in enumerate(Doc_date, start=4):  # i 對應 Excel 的行號（從 1 開始）
    wc.cell(row=i, column=1, value=item)  # 第一列（A欄）填入資料
    wc.cell(row=i, column=2, value='車用柴油')  # 第二列（B欄）填入資料
# 資料開始寫入
row = 4  # 從第四行開始寫入資料
for information in data3:
    for col_idx, value in enumerate(information.values(), start=3):  # 從 column=3 開始
        wc.cell(row=row, column=col_idx, value=value)
    row += 1  # 移動到下一行
# 套用黑框
for row in wc.iter_rows(min_row=3, max_row=len(Doc_date)+3, min_col=1, max_col=len(headers2)):
    for cell in row:
        cell.border = black_border
# 自動調整欄寬
for col_idx, col_cells in enumerate(wc.columns, start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)  # 取得欄位字母 (如 A, B, C...)
    for cell in col_cells:
        if cell.value:
            # 取得最大字元長度
            max_length = max(max_length, len(str(cell.value)))
    # 預估調整寬度（+2 以增加間距）
    adjusted_width = (max_length + 4) * 1.2  # 乘以 1.2 來更準確反映 Excel 的寬度
    wc.column_dimensions[col_letter].width = adjusted_width  # 設定欄位寬度


# 類別一-滅火器
wz.create_sheet(sheet[2])  # 插入工作表3 在最後方
wd = wz[sheet[2]] 
wd.title = sheet[2]
# 資料
data4 = [
    {
        'item_name': 'AT555',
        'ingredient': '乾粉',
        'specification': '公升',
        'unit': '1,000',
        'usage': '1,500',
        'remark': 'none', 
     },
     {
        'item_name': '乾粉2',
        'ingredient': '',
        'unit': '',
        'money': '',
        'liters': '',
        'remark': '', 
     },  
     ]
# 定義標題
# 合併 A1 到 F1 儲存格
wd.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
merged_cell_3 = wd.cell(row=1, column=1, value=sheet[2])  # 設定合併後的內容
merged_cell_3.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
merged_cell_3.fill = PatternFill(fill_type="solid", fgColor="ffffcc00") # 設定儲存格的背景樣式
# 第二行
wd.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
merged_cell = wd.cell(row=2, column=1, value='公司本年度二氧化碳滅火器之請購及換藥紀錄（無則免填）')  # 設定合併後的內容
# 第三行
headers2 = ['品名', '成分', '規格(重量)', '單位', '使用量(支)', '備註']
for col_idx, header in enumerate(headers2, start=1):
    cell2 = wd.cell(row=3, column=col_idx, value=header)
    cell2.fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD") # 設定儲存格的背景樣式
# 資料開始寫入
row = 4  # 從第四行開始寫入資料
for information in data4:
    for col_idx, value in enumerate(information.values(), start=1):  # 從 column=1 開始
        wd.cell(row=row, column=col_idx, value=value)
    row += 1  # 移動到下一行
# 欄位備註
Note = ['文字', '選擇', '數字', '選擇', '數字', '文字']
for col_idx, N in enumerate(Note, start=1):
    wd.cell(row=len(data4)+4, column=col_idx, value=N)
# 套用黑框
for row in wd.iter_rows(min_row=3, max_row=len(data4)+3, min_col=1, max_col=len(headers2)):
    for cell in row:
        cell.border = black_border
# 自動調整欄寬
for col_idx, col_cells in enumerate(wd.columns, start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)  # 取得欄位字母 (如 A, B, C...)
    for cell in col_cells:
        if cell.value:
            # 取得最大字元長度
            max_length = max(max_length, len(str(cell.value)))
    # 預估調整寬度（+2 以增加間距）
    adjusted_width = (max_length + 4) * 1.2  # 乘以 1.2 來更準確反映 Excel 的寬度
    wd.column_dimensions[col_letter].width = adjusted_width  # 設定欄位寬度


# 類別一-工作時數(員工)
wz.create_sheet(sheet[3]) 
we = wz[sheet[3]] 
we.title = sheet[3]
# 資料
data5 = [
    {
        # 'period_date': '1月',
        'employee_number': 23,
        'daily_hours': 8,
        'workday': 21,
        'overtime': 0,
        'sick_leave': 0, 
        'personal_leave': 0,
        'business_trip': 0,
        'wedding_and_funeral': 0,
        'special_leave': 0,
        'remark': 'none',
     },
     {
        # 'period_date': '2月',
        'employee_number': 23,
        'daily_hours': 8,
        'workday': 15,
        'overtime': 0,
        'sick_leave': 0, 
        'personal_leave': 0,
        'business_trip': 0,
        'wedding_and_funeral': 0,
        'special_leave': 0,
        'remark': 'none',
     },
     {
        # 'period_date': '3月',
        'employee_number': 22,
        'daily_hours': 8,
        'workday': 23,
        'overtime': 0,
        'sick_leave': 0, 
        'personal_leave': 0,
        'business_trip': 0,
        'wedding_and_funeral': 0,
        'special_leave': 0,
        'remark': 'none',
     },
     {
        # 'period_date': '4月',
        'employee_number': 22,
        'daily_hours': 8,
        'workday': 19,
        'overtime': 0,
        'sick_leave': 0, 
        'personal_leave': 0,
        'business_trip': 0,
        'wedding_and_funeral': 0,
        'special_leave': 0,
        'remark': 'none',
     }, 
     {
        # 'period_date': '5月',
        'employee_number': 24,
        'daily_hours': 8,
        'workday': 21,
        'overtime': 0,
        'sick_leave': 0, 
        'personal_leave': 0,
        'business_trip': 0,
        'wedding_and_funeral': 0,
        'special_leave': 0,
        'remark': 'none',
     },
     {
        # 'period_date': '6月',
        'employee_number': 23,
        'daily_hours': 8,
        'workday': 21,
        'overtime': 0,
        'sick_leave': 0, 
        'personal_leave': 0,
        'business_trip': 0,
        'wedding_and_funeral': 0,
        'special_leave': 0,
        'remark': 'none',
     },
     {
        # 'period_date': '7月',
        'employee_number': 24,
        'daily_hours': 8,
        'workday': 20,
        'overtime': 0,
        'sick_leave': 0, 
        'personal_leave': 0,
        'business_trip': 0,
        'wedding_and_funeral': 0,
        'special_leave': 0,
        'remark': 'none',
     },
     {
        # 'period_date': '8月',
        'employee_number': 24,
        'daily_hours': 8,
        'workday': 23,
        'overtime': 0,
        'sick_leave': 0, 
        'personal_leave': 0,
        'business_trip': 0,
        'wedding_and_funeral': 0,
        'special_leave': 0,
        'remark': 'none',
     },
     {
        # 'period_date': '9月',
        'employee_number': 23,
        'daily_hours': 8,
        'workday': 21,
        'overtime': 0,
        'sick_leave': 0, 
        'personal_leave': 0,
        'business_trip': 0,
        'wedding_and_funeral': 0,
        'special_leave': 0,
        'remark': 'none',
     },
     {
        # 'period_date': '10月',
        'employee_number': 23,
        'daily_hours': 8,
        'workday': 20,
        'overtime': 0,
        'sick_leave': 0, 
        'personal_leave': 0,
        'business_trip': 0,
        'wedding_and_funeral': 0,
        'special_leave': 0,
        'remark': 'none',
     },
     {
        # 'period_date': '11月',
        'employee_number': 25,
        'daily_hours': 8,
        'workday': 22,
        'overtime': 0,
        'sick_leave': 0, 
        'personal_leave': 0,
        'business_trip': 0,
        'wedding_and_funeral': 0,
        'special_leave': 0,
        'remark': 'none',
     },
     {
        # 'period_date': '12月',
        'employee_number': 25,
        'daily_hours': 8,
        'workday': 22,
        'overtime': 0,
        'sick_leave': 0, 
        'personal_leave': 0,
        'business_trip': 0,
        'wedding_and_funeral': 0,
        'special_leave': 0,
        'remark': 'none',
     },
     ]
# 定義標題
# 合併儲存格
we.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
merged_cell_4 = we.cell(row=1, column=1, value="員工平均工作時數")  # 設定合併後的內容
merged_cell_4.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
merged_cell_4.fill = PatternFill(fill_type="solid", fgColor="ffffcc00") # 設定儲存格的背景樣式
# 第二行
headers3 = ['月份', '員工數', '每日工時', '每月工作日數', '總加班時數', '總病假時數', '總事假時數', '總出差時數', '總婚喪時數', '總特休時數', '備註']
for col_idx, header in enumerate(headers3, start=1):
    cell2 = we.cell(row=2, column=col_idx, value=header)
    cell2.fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD") # 設定儲存格的背景樣式
# 月份
period_date = ['1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月','總計']
for i, item in enumerate(period_date, start=3):  # i 對應 Excel 的行號（從 3 開始）
    we.cell(row=i, column=1, value=item)  # 第一列（A欄）填入資料
# 資料開始寫入
row = 3  # 從第三行開始寫入資料
for information in data5:
    for col_idx, value in enumerate(information.values(), start=2):  # 從 column=2 開始
        we.cell(row=row, column=col_idx, value=value)
    row += 1  # 移動到下一行
# 設定總計行的 SUM 公式 (假設員工數在 B 欄、每月工作日數在 D 欄)
sum_row = len(period_date) + 2  # 總計的行號 (最後一個月份 + 1)
we.cell(row=sum_row, column=2, value="=SUM(B3:B14)")  # B 欄的總計 (員工數)
we.cell(row=sum_row, column=4, value="=SUM(D3:D14)")  # D 欄的總計 (每月工作日數)
# 欄位備註
Note1 = ['', '', '不含休息時間', '', '可不填', '可不填', '可不填', '可不填', '可不填', '可不填', '可不填']
for col_idx, N in enumerate(Note1, start=1):
    we.cell(row=len(period_date)+2, column=col_idx, value=N)
# 套用黑框
for row in we.iter_rows(min_row=2, max_row=len(period_date)+1, min_col=1, max_col=len(headers3)):
    for cell in row:
        cell.border = black_border
# 自動調整欄寬
for col_idx, col_cells in enumerate(we.columns, start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)  # 取得欄位字母 (如 A, B, C...)
    for cell in col_cells:
        if cell.value:
            # 取得最大字元長度
            max_length = max(max_length, len(str(cell.value)))
    # 預估調整寬度（+2 以增加間距）
    adjusted_width = (max_length + 6) * 1.4  # 乘以 1.2 來更準確反映 Excel 的寬度
    we.column_dimensions[col_letter].width = adjusted_width  # 設定欄位寬度

# 類別一-工作時數(非員工)
wz.create_sheet(sheet[4]) 
wf = wz[sheet[4]] 
wf.title = sheet[4]
# 資料
data6 = [
    {
        # 'period_date': '1月',
        'nonemployee_number': 23,
        'total_hours': 8,
        'total_days': 21,
        'remark': 'none',
     },
     {
        # 'period_date': '2月',
        'nonemployee_number': 23,
        'total_hours': 8,
        'total_days': 15,
        'remark': 'none',
     },
     {
        # 'period_date': '3月',
        'nonemployee_number': 22,
        'total_hours': 8,
        'total_days': 23,
        'remark': 'none',
     },
     {
        # 'period_date': '4月',
        'nonemployee_number': 22,
        'total_hours': 8,
        'total_days': 19,
        'remark': 'none',
     }, 
     {
        # 'period_date': '5月',
        'nonemployee_number': 24,
        'total_hours': 8,
        'total_days': 21,
        'remark': 'none',
     },
     {
        # 'period_date': '6月',
        'nonemployee_number': 23,
        'total_hours': 8,
        'total_days': 21,
        'remark': 'none',
     },
     {
        # 'period_date': '7月',
        'nonemployee_number': 24,
        'total_hours': 8,
        'total_days': 20,
        'remark': 'none',
     },
     {
        # 'period_date': '8月',
        'nonemployee_number': 24,
        'total_hours': 8,
        'total_days': 23,
        'remark': 'none',
     },
     {
        # 'period_date': '9月',
        'nonemployee_number': 23,
        'total_hours': 8,
        'total_days': 21,
        'remark': 'none',
     },
     {
        # 'period_date': '10月',
        'nonemployee_number': 23,
        'total_hours': 8,
        'total_days': 20,
        'remark': 'none',
     },
     {
        # 'period_date': '11月',
        'nonemployee_number': 25,
        'total_hours': 8,
        'total_days': 22,
        'remark': 'none',
     },
     {
        # 'period_date': '12月',
        'nonemployee_number': 25,
        'total_hours': 8,
        'total_days': 22,
        'remark': 'none',
     },
     ]
# 定義標題
# 合併儲存格
wf.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
merged_cell_5 = wf.cell(row=1, column=1, value=sheet[4])  # 設定合併後的內容
merged_cell_5.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
merged_cell_5.fill = PatternFill(fill_type="solid", fgColor="ffffcc00") # 設定儲存格的背景樣式
# 第二行
headers4 = ['月份', '人數', '總工時時數', '總工作人天', '備註']
for col_idx, header in enumerate(headers4, start=1):
    cell2 = wf.cell(row=2, column=col_idx, value=header)
    cell2.fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD") # 設定儲存格的背景樣式
# 月份
period_date2 = ['1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']
for i, item in enumerate(period_date2, start=3):  # i 對應 Excel 的行號（從 3 開始）
    wf.cell(row=i, column=1, value=item)  # 第一列（A欄）填入資料
# 資料開始寫入
row = 3  # 從第三行開始寫入資料
for information in data6:
    for col_idx, value in enumerate(information.values(), start=2):  # 從 column=2 開始
        wf.cell(row=row, column=col_idx, value=value)
    row += 1  # 移動到下一行
# 欄位備註
Note = ['', '數字', '數字', '數字', '文字(可不填寫)']
for col_idx, N in enumerate(Note, start=1):
    wf.cell(row=len(period_date2)+3, column=col_idx, value=N)
# 套用黑框
for row in wf.iter_rows(min_row=2, max_row=len(period_date2)+2, min_col=1, max_col=len(headers4)):
    for cell in row:
        cell.border = black_border
# 自動調整欄寬
for col_idx, col_cells in enumerate(wf.columns, start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)  # 取得欄位字母 (如 A, B, C...)
    for cell in col_cells:
        if cell.value:
            # 取得最大字元長度
            max_length = max(max_length, len(str(cell.value)))
    # 預估調整寬度（+2 以增加間距）
    adjusted_width = (max_length + 6) * 1.4  # 乘以 1.2 來更準確反映 Excel 的寬度
    wf.column_dimensions[col_letter].width = adjusted_width  # 設定欄位寬度


# 類別一-冷媒
wz.create_sheet(sheet[5])
wg = wz[sheet[5]] 
wg.title = sheet[5]
# 資料
data7 = [
    {
        'device_type': '請選擇設備類型',
        'location': '',
        'refrigerant_type': '選擇冷媒類型',
        'unit': '選擇單位',
        'usage_unit': 0,
        'usage': 0, 
        'escape_rate': 0, 
        'remark': 0, 
     },
    {
        'device_type': '請選擇設備類型',
        'location': '',
        'refrigerant_type': '選擇冷媒類型',
        'unit': '選擇單位',
        'usage_unit': 0,
        'usage': 0, 
        'escape_rate': 0, 
        'remark': 0, 
     },
    {
        'device_type': '請選擇設備類型',
        'location': '',
        'refrigerant_type': '選擇冷媒類型',
        'unit': '選擇單位',
        'usage_unit': 0,
        'usage': 0, 
        'escape_rate': 0, 
        'remark': 0, 
     },
    {
        'device_type': '請選擇設備類型',
        'location': '',
        'refrigerant_type': '選擇冷媒類型',
        'unit': '選擇單位',
        'usage_unit': 0,
        'usage': 0, 
        'escape_rate': 0, 
        'remark': 0, 
     },
    {
        'device_type': '請選擇設備類型',
        'location': '',
        'refrigerant_type': '選擇冷媒類型',
        'unit': '選擇單位',
        'usage_unit': 0,
        'usage': 0, 
        'escape_rate': 0, 
        'remark': 0, 
     },
     ]
# 定義標題
# 合併儲存格
wg.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
merged_cell_6 = wg.cell(row=1, column=1, value='公司冷媒清冊－冰箱、冷氣機、飲水機、冰水主機、空壓機')  # 設定合併後的內容
merged_cell_6.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
merged_cell_6.fill = PatternFill(fill_type="solid", fgColor="ffffcc00") # 設定儲存格的背景樣式
# 第二行
wg.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
merged_cell_6_2 = wg.cell(row=2, column=1, value='檢視往年資料')  # 設定合併後的內容
merged_cell_6_2.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
# 第三行
headers5 = ['設備類型', '設備位置', '冷媒類型', '填充量單位', '填充量', '數量', '逸散率(%)', '備註']
for col_idx, header in enumerate(headers5, start=1):
    cell2 = wg.cell(row=3, column=col_idx, value=header)
    cell2.fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD") # 設定儲存格的背景樣式
# 資料開始寫入
row = 4  # 從第四行開始寫入資料
for information in data7:
    for col_idx, value in enumerate(information.values(), start=1):  # 從 column=1 開始
        wg.cell(row=row, column=col_idx, value=value)
    row += 1  # 移動到下一行
# 欄位備註
Note2 = ['選擇','文字','選擇','選擇', '數字', '數字', '數字(可不填)', '可不填']
for col_idx, N in enumerate(Note2, start=1):
    wg.cell(row=len(data7)+4, column=col_idx, value=N)
# 套用黑框
for row in wg.iter_rows(min_row=3, max_row=len(data7)+3, min_col=1, max_col=len(headers5)):
    for cell in row:
        cell.border = black_border
# 自動調整欄寬
for col_idx, col_cells in enumerate(wg.columns, start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)  # 取得欄位字母 (如 A, B, C...)
    for cell in col_cells:
        if cell.value:
            # 取得最大字元長度
            max_length = max(max_length, len(str(cell.value)))
    # 預估調整寬度（+2 以增加間距）
    adjusted_width = (max_length + 6) * 1.4  # 乘以 1.2 來更準確反映 Excel 的寬度
    wg.column_dimensions[col_letter].width = adjusted_width  # 設定欄位寬度


# 類別一-固定式燃燒
wz.create_sheet(sheet[6])
wh = wz[sheet[6]] 
wh.title = sheet[6]
# 資料
data8 = [
    {
        'device_type': '',
        'energy_type': '選擇能源類型',
        'time': '選擇時期',
        'usage': 0,
        'unit': '選擇單位',
        'remark': 'none', 
     },
    {
        'device_type': '',
        'energy_type': '選擇能源類型',
        'time': '選擇時期',
        'usage': 0,
        'unit': '選擇單位',
        'remark': 'none', 
     },
    {
        'device_type': '',
        'energy_type': '選擇能源類型',
        'time': '選擇時期',
        'usage': 0,
        'unit': '選擇單位',
        'remark': 'none', 
     },
     ]
# 定義標題
# 合併儲存格
wh.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
merged_cell_7 = wh.cell(row=1, column=1, value='公司固定式燃燒清冊')  # 設定合併後的內容
merged_cell_7.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
merged_cell_7.fill = PatternFill(fill_type="solid", fgColor="ffffcc00") # 設定儲存格的背景樣式
# 第二行
wh.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
merged_cell_7_2 = wh.cell(row=2, column=1, value='檢視往年資料')  # 設定合併後的內容
merged_cell_7_2.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
# 第三行
headers6 = ['設備類型', '能源類型', '時期', '使用量', '單位', '備註']
for col_idx, header in enumerate(headers6, start=1):
    cell2 = wh.cell(row=3, column=col_idx, value=header)
    cell2.fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD") # 設定儲存格的背景樣式
# 資料開始寫入
row = 4  # 從第四行開始寫入資料
for information in data8:
    for col_idx, value in enumerate(information.values(), start=1):  # 從 column=1 開始
        wh.cell(row=row, column=col_idx, value=value)
    row += 1  # 移動到下一行
# 欄位備註
Note3 = ['文字','選擇','', '數字', '選擇', '可不填']
for col_idx, N in enumerate(Note3, start=1):
    wh.cell(row=len(data8)+4, column=col_idx, value=N)
# 套用黑框
for row in wh.iter_rows(min_row=3, max_row=len(data8)+3, min_col=1, max_col=len(headers6)):
    for cell in row:
        cell.border = black_border
# 自動調整欄寬
for col_idx, col_cells in enumerate(wh.columns, start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)  # 取得欄位字母 (如 A, B, C...)
    for cell in col_cells:
        if cell.value:
            # 取得最大字元長度
            max_length = max(max_length, len(str(cell.value)))
    # 預估調整寬度（+2 以增加間距）
    adjusted_width = (max_length + 6) * 1.4  # 乘以 1.2 來更準確反映 Excel 的寬度
    wh.column_dimensions[col_letter].width = adjusted_width  # 設定欄位寬度


# 類別一-產生溫室氣體排放製程
wz.create_sheet(sheet[7])
wI = wz[sheet[7]] 
wI.title = sheet[7]
# 資料
data9 = [
    {
        'date': '選擇日期',
        'fuel_category': '選擇原料',
        'usage': 0,
        'unit': '選擇單位',
        'remark': 'none', 
     },
    {
        'date': '選擇日期',
        'fuel_category': '選擇原料',
        'usage': 0,
        'unit': '選擇單位',
        'remark': 'none',
     },
    {
        'date': '選擇日期',
        'fuel_category': '選擇原料',
        'usage': 0,
        'unit': '選擇單位',
        'remark': 'none', 
     },
     ]
# 定義標題
# 合併儲存格
wI.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
merged_cell_8 = wI.cell(row=1, column=1, value='公司產生溫室氣體排放製程調查')  # 設定合併後的內容
merged_cell_8.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
merged_cell_8.fill = PatternFill(fill_type="solid", fgColor="ffffcc00") # 設定儲存格的背景樣式
# 第二行
wI.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
merged_cell_8_2 = wI.cell(row=2, column=1, value='檢視往年資料')  # 設定合併後的內容
merged_cell_8_2.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
# 第三行
headers7 = ['日期', '使用原料', '使用量', '單位', '備註']
for col_idx, header in enumerate(headers7, start=1):
    cell2 = wI.cell(row=3, column=col_idx, value=header)
    cell2.fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD") # 設定儲存格的背景樣式
# 資料開始寫入
row = 4  # 從第四行開始寫入資料
for information in data9:
    for col_idx, value in enumerate(information.values(), start=1):  # 從 column=1 開始
        wI.cell(row=row, column=col_idx, value=value)
    row += 1  # 移動到下一行
# 欄位備註
Note4 = ['日期','選擇', '數字', '選擇', '可不填']
for col_idx, N in enumerate(Note4, start=1):
    wI.cell(row=len(data9)+4, column=col_idx, value=N)
# 套用黑框
for row in wI.iter_rows(min_row=3, max_row=len(data9)+3, min_col=1, max_col=len(headers7)):
    for cell in row:
        cell.border = black_border
# 自動調整欄寬
for col_idx, col_cells in enumerate(wI.columns, start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)  # 取得欄位字母 (如 A, B, C...)
    for cell in col_cells:
        if cell.value:
            # 取得最大字元長度
            max_length = max(max_length, len(str(cell.value)))
    # 預估調整寬度（+2 以增加間距）
    adjusted_width = (max_length + 6) * 1.4  # 乘以 1.2 來更準確反映 Excel 的寬度
    wI.column_dimensions[col_letter].width = adjusted_width  # 設定欄位寬度


# 類別一-廠內機具
wz.create_sheet(sheet[8])
wJ = wz[sheet[8]] 
wJ.title = sheet[8]
# 資料
data10 = [
    {
        'machinery_location':'',
        'energy_type': '選擇能源類型',
        'month': '選擇時期',
        'usage': 0,
        'unit': '選擇單位',
        'remark': 'none',  
     },
    {
        'machinery_location':'',
        'energy_type': '選擇能源類型',
        'month': '選擇時期',
        'usage': 0,
        'unit': '選擇單位',
        'remark': 'none', 
     },
    {
        'machinery_location':'',
        'energy_type': '選擇能源類型',
        'month': '選擇時期',
        'usage': 0,
        'unit': '選擇單位',
        'remark': 'none', 
     },
     ]
# 定義標題
# 合併儲存格
wJ.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
merged_cell_9 = wJ.cell(row=1, column=1, value='廠內機具高機清冊')  # 設定合併後的內容
merged_cell_9.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
merged_cell_9.fill = PatternFill(fill_type="solid", fgColor="ffffcc00") # 設定儲存格的背景樣式
# 第二行
wJ.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
merged_cell_9_2 = wJ.cell(row=2, column=1, value='檢視往年資料')  # 設定合併後的內容
merged_cell_9_2.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
# 第三行
headers8 = ['設備位置', '能源類型', '月份', '使用量', '單位', '備註']
for col_idx, header in enumerate(headers8, start=1):
    cell2 = wJ.cell(row=3, column=col_idx, value=header)
    cell2.fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD") # 設定儲存格的背景樣式
# 資料開始寫入
row = 4  # 從第四行開始寫入資料
for information in data10:
    for col_idx, value in enumerate(information.values(), start=1):  # 從 column=1 開始
        wJ.cell(row=row, column=col_idx, value=value)
    row += 1  # 移動到下一行
# 欄位備註
Note5 = ['文字','選擇', '', '數字', '選擇', '可不填']
for col_idx, N in enumerate(Note5, start=1):
    wJ.cell(row=len(data10)+4, column=col_idx, value=N)
# 套用黑框
for row in wJ.iter_rows(min_row=3, max_row=len(data10)+3, min_col=1, max_col=len(headers8)):
    for cell in row:
        cell.border = black_border
# 自動調整欄寬
for col_idx, col_cells in enumerate(wJ.columns, start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)  # 取得欄位字母 (如 A, B, C...)
    for cell in col_cells:
        if cell.value:
            # 取得最大字元長度
            max_length = max(max_length, len(str(cell.value)))
    # 預估調整寬度（+2 以增加間距）
    adjusted_width = (max_length + 6) * 1.4  # 乘以 1.2 來更準確反映 Excel 的寬度
    wJ.column_dimensions[col_letter].width = adjusted_width  # 設定欄位寬度


# 類別一-緊急發電機
wz.create_sheet(sheet[9])
wK = wz[sheet[9]] 
wK.title = sheet[9]
# 資料
data11 = [
    {
        'Doc_date':'請輸入西元/月/日',
        'Doc_number': '',
        'usage': 0,
        'remark': 'none',  
     },
    {
        'Doc_date':'請輸入西元/月/日',
        'Doc_number': '',
        'usage': 0,
        'remark': 'none', 
     },
    {
        'Doc_date':'請輸入西元/月/日',
        'Doc_number': '',
        'usage': 0,
        'remark': 'none', 
     },
     ]
# 定義標題
# 合併儲存格
wK.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
merged_cell_10 = wK.cell(row=1, column=1, value='緊急發電機柴油填充紀錄')  # 設定合併後的內容
merged_cell_10.alignment = Alignment(horizontal="center", vertical="center")  # 文字置中
merged_cell_10.fill = PatternFill(fill_type="solid", fgColor="ffffcc00") # 設定儲存格的背景樣式
merged_cell_10.border = black_border  # 套用黑色邊框
# 第二行
headers9 = ['發票日期', '發票號碼', '使用量(公升)', '備註']
for col_idx, header in enumerate(headers9, start=1):
    cell2 = wK.cell(row=2, column=col_idx, value=header)
    cell2.fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD") # 設定儲存格的背景樣式
    cell2.border = black_border  # 套用黑色邊框
# 資料開始寫入
row = 3  # 從第三行開始寫入資料
for information in data11:
    for col_idx, value in enumerate(information.values(), start=1):  # 從 column=1 開始
        data11_cell = wK.cell(row=row, column=col_idx, value=value)
        data11_cell.border = black_border  # 套用黑色邊框
    row += 1  # 移動到下一行
# 欄位備註
Note6 = ['日期','', '數字', '文字']
for col_idx, N in enumerate(Note6, start=1):
    wK.cell(row=len(data11)+3, column=col_idx, value=N)
# 套用黑框
for row in wK.iter_rows(min_row=3, max_row=len(data11)+2, min_col=1, max_col=len(headers9)):
    for cell in row:
        cell.border = black_border
# 自動調整欄寬
for col_idx, col_cells in enumerate(wK.columns, start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)  # 取得欄位字母 (如 A, B, C...)
    for cell in col_cells:
        if cell.value:
            # 取得最大字元長度
            max_length = max(max_length, len(str(cell.value)))
    # 預估調整寬度（+2 以增加間距）
    adjusted_width = (max_length + 6) * 1.4  # 乘以 1.2 來更準確反映 Excel 的寬度
    wK.column_dimensions[col_letter].width = adjusted_width  # 設定欄位寬度




# 整個excel表_存檔
wz.save('碳盤查基準活動數據.xlsx')