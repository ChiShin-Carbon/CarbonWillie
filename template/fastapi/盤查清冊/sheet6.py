from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import style  # 引入樣式模組

def create_refrigerant_sheet(wb):
    """建立 '類別一-冷媒' 工作表"""
    sheet_name = "類別一-冷媒"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = ws.cell(row=1, column=1, value="公司冷媒清冊－冰箱、冷氣機、飲水機、冰水主機、空壓機")
    title_cell.alignment = style.center_alignment
    title_cell.fill = style.yellow_fill  # 標題背景色

    # 第二行：說明文字
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.cell(row=2, column=1, value="檢視往年資料").alignment = style.center_alignment

    # 第三行：表頭
    headers = ['設備類型', '設備位置', '冷媒類型', '填充量單位', '填充量', '數量', '逸散率(%)', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 灰色背景

    # 填入數據
    data = [
        {'device_type': '請選擇設備類型', 'location': '', 'refrigerant_type': '選擇冷媒類型',
         'unit': '選擇單位', 'usage_unit': 0, 'usage': 0, 'escape_rate': 0, 'remark': 0}
    ] * 5  # 建立 5 行預設數據

    row = 4
    for item in data:
        for col_idx, key in enumerate(['device_type', 'location', 'refrigerant_type', 'unit',
                                       'usage_unit', 'usage', 'escape_rate', 'remark'], start=1):
            ws.cell(row=row, column=col_idx, value=item.get(key, ''))
        row += 1

    # 第 9 行：欄位說明
    notes = ['選擇', '文字', '選擇', '選擇', '數字', '數字', '數字(可不填)', '可不填']
    for col_idx, note in enumerate(notes, start=1):
        ws.cell(row=row, column=col_idx, value=note)

    # 套用邊框
    for row_cells in ws.iter_rows(min_row=3, max_row=len(data)+3, min_col=1, max_col=len(headers)):
        for cell in row_cells:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2
