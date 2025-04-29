from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import requests
import style  # 引入樣式模組

# 設定年份
year = 2024
api_url = f"http://localhost:8000/employee_data_by_year/{year}"  # 依實際 API 位址修改

def create_work_hours_sheet(wb, data):
    sheet_name = "類別一-工作時數(員工)"
    ws = wb.create_sheet(title=sheet_name)
    
    # 設定標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    merged_cell = ws.cell(row=1, column=1, value="員工平均工作時數")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill  # 黃色背景
    
    # 第二行：標題列
    headers = ['月份', '員工數', '每日工時', '每月工作日數', '總加班時數', '總病假時數', '總事假時數', '總出差時數', '總婚喪時數', '總特休時數', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 灰色背景
    
    # 使用 API 資料填充
    month_map = {"01": "1月", "02": "2月", "03": "3月", "04": "4月", "05": "5月", "06": "6月",
                "07": "7月", "08": "8月", "09": "9月", "10": "10月", "11": "11月", "12": "12月"}
    
    # 按月份排序資料
    sorted_data = sorted(data, key=lambda x: x['month'])
    
    for i, item in enumerate(sorted_data, start=3):
        # 處理月份顯示格式
        month_display = month_map.get(item['month'], f"{item['month']}月")
        
        ws.cell(row=i, column=1, value=month_display)  # A 欄：月份
        ws.cell(row=i, column=2, value=item['employee_number'])  # B 欄：員工數
        ws.cell(row=i, column=3, value=item['daily_hours'])  # C 欄：每日工時
        ws.cell(row=i, column=4, value=item['workday'])  # D 欄：每月工作日數
        ws.cell(row=i, column=5, value=item.get('overtime', ''))  # E 欄：總加班時數
        ws.cell(row=i, column=6, value=item.get('sick_leave', ''))  # F 欄：總病假時數
        ws.cell(row=i, column=7, value=item.get('personal_leave', ''))  # G 欄：總事假時數
        ws.cell(row=i, column=8, value=item.get('business_trip', ''))  # H 欄：總出差時數
        ws.cell(row=i, column=9, value=item.get('wedding_and_funeral', ''))  # I 欄：總婚喪時數
        ws.cell(row=i, column=10, value=item.get('special_leave', ''))  # J 欄：總特休時數
        ws.cell(row=i, column=11, value=item.get('remark', ''))  # K 欄：備註
    
    # 設定總計行
    total_row = len(sorted_data) + 3
    ws.cell(row=total_row, column=1, value="總計")
    ws.cell(row=total_row, column=2, value=f"=SUM(B3:B{total_row-1})")
    ws.cell(row=total_row, column=4, value=f"=SUM(D3:D{total_row-1})")
    ws.cell(row=total_row, column=5, value=f"=SUM(E3:E{total_row-1})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F3:F{total_row-1})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G3:G{total_row-1})")
    ws.cell(row=total_row, column=8, value=f"=SUM(H3:H{total_row-1})")
    ws.cell(row=total_row, column=9, value=f"=SUM(I3:I{total_row-1})")
    ws.cell(row=total_row, column=10, value=f"=SUM(J3:J{total_row-1})")
    
    # 套用框線
    for row in ws.iter_rows(min_row=2, max_row=total_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border
    
    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2


# 主程式區塊
def main():
    try:
        response = requests.get(api_url)
        
        if response.status_code == 200:
            data = response.json()
            
            wb = Workbook()
            wb.remove(wb.active)  # 移除預設工作表
            create_work_hours_sheet(wb, data)
            
            output_filename = f"employee_hours_{year}.xlsx"
            wb.save(output_filename)
            print(f"Excel 檔案已生成: {output_filename}")
        else:
            print(f"API 請求失敗，狀態碼: {response.status_code}，錯誤訊息: {response.text}")
    except Exception as e:
        print(f"執行發生錯誤: {e}")

if __name__ == "__main__":
    main()