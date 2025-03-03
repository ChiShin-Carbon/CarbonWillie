from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import style  # 引入樣式模組

def create_sheetdevision(wb):
    sheet_name = "分工說明表"
    ws = wb.create_sheet(title=sheet_name)
    # 定義標題
    headers = ['表單名稱', '部門', '姓名(主要填寫人)', '電子郵件', '電話及分機號碼', '平台帳號', '平台密碼', '平台權限', '備註']
    
    # 設定標題樣式
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = style.center_alignment
        cell.fill = style.yellow_fill

    # 各表單名稱
    sheet_names = [
        '類別一-公務車(汽油)', '類別一-公務車(柴油)', '類別一-滅火器',
        '類別一-工作時數(員工)', '類別一-工作時數(非員工)', '類別一-冷媒',
        '類別一-固定式', '類別一-產生溫室氣體的排放製程', '類別一-廠內機具',
        '類別一-緊急發電機', '類別一-焊條', '類別一-氣體斷路器(GCB)',
        '類別一-其他', '類別二-電力使用量', '類別二-間接蒸氣(汽電共生廠有做溫室氣體盤查)',
        '類別二-間接蒸氣(汽電共生廠沒有做溫室氣體盤查)', '類別二-其他'
    ]

    # 填充表單名稱
    for i, name in enumerate(sheet_names, start=2):
        ws.cell(row=i, column=1, value=name)

    # 預設人員資料
    data = [
        {
            'department': '業務部', 'name': '陳欣柔', 'email': 'chen@gmail.com',
            'phone': '09722222', 'address': 'admin', 'password': 'chen',
            'role': 'admin', 'remark': 'none'
        }
    ] + [{}] * (len(sheet_names) - 1)  # 其餘空白行

    # 寫入資料
    for row_idx, person in enumerate(data, start=2):
        for col_idx, key in enumerate(['department', 'name', 'email', 'phone', 'address', 'password', 'role', 'remark'], start=2):
            ws.cell(row=row_idx, column=col_idx, value=person.get(key, ''))

    # 最後：說明文字
    ws.merge_cells(start_row=len(data)+4, start_column=1, end_row=len(data) + 4, end_column=9)
    merged_cell_explain = ws.cell(row=len(data)+4, column=1, value="平台權限:\n"
            "1. 廠商負責人: 負責控管整個專案，可以查看及編輯所有填寫者的回覆。\n"
            "2. 主管階級: 檢視及編輯專案內的填寫資料。\n"
            "   *主管階級預設只能檢視專案內容，若需新增編輯權限，請在備註說明。*\n"
            "3. 一般使用者: 填寫自己所被分配的類別，如環安被分配到填寫冷媒，\n"
            "   那只能看到冷媒填寫頁面，不能查看其他類別 (如通勤、差旅) 之填寫頁面。"
    )
    merged_cell_explain.alignment = style.newline # 自動換行
    merged_cell_explain.fill = style.yellow_fill  # 背景色

    # 套用黑色邊框
    for row in ws.iter_rows(min_row=1, max_row=len(sheet_names) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border
    # # 設定較小的欄寬
    # col_letter = get_column_letter(1)  # 取得第1欄的字母標記（通常是 'A'）
    # ws.column_dimensions[col_letter].width = 30  # 調整寬度（可根據需求修改）

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        adjusted_width = (max_length + 4) * 1.3  # 調整寬度
        if col_idx == 1:  # 第一欄加長
            adjusted_width = 50
        elif col_idx == len(data)+4:  # 第一欄加長
            for row_idx in range(len(data) + 4, len(data) + 5):
                ws.row_dimensions[row_idx].height = 6000  # 設定較高的行高
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
