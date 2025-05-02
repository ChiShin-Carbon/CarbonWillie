from openpyxl import Workbook
import sheetdevision
import sheet1
import sheet2
import sheet3
import sheet4
import sheet5
import sheet6
import sheet7
import sheet8
import sheet9
import sheet10
import sheet11
import sheet12
import sheet13
import sheet14
import sheet15
import sheet16
import sheet17
import sheet18

def main():
    wb = Workbook()
    wb.remove(wb.active)  # 移除預設的工作表
    
    # 建立各個工作表
    # sheetdevision.create_sheetdevision(wb)   #分工說明表
    # sheet1.create_sheet1(wb)   #類別一-公務車(汽油)
    # sheet2.create_sheet2(wb)   #類別一-公務車(柴油)
    # sheet3.create_fire_extinguisher_sheet(wb)   #類別一-滅火器
    # sheet4.create_work_hours_sheet(wb)   #類別一-工作時數(員工)
    # sheet5.create_nonemployee_work_hours_sheet(wb)   #類別一-工作時數(非員工)
    # sheet6.create_refrigerant_sheet(wb)   #類別一-冷煤
    # sheet7.create_fixed_combustion_sheet(wb)   #類別一-固定式燃燒
    # sheet8.create_greenhouse_gas_sheet(wb)   #類別一-產生溫室氣體排放製程
    # sheet9.create_internal_machinery_sheet(wb)   #類別一-廠內機具
    # sheet10.create_emergency_generator_sheet(wb)   #類別一-緊急發電機
    # sheet11.create_welding_rod_sheet(wb)   #類別一-焊條
    # sheet12.create_gcb_sheet(wb)   #類別一-氣體斷路器(GCB)
    # sheet13.create_other_category_sheet(wb)   #類別一-其他
    # sheet14.create_power_usage_sheet(wb)   #類別二-電力使用量
    # sheet15.create_indirect_steam_sheet(wb)   #類別二-間接蒸氣(汽電共生廠有做溫室氣體盤查)
    # sheet16.create_non_indirect_steam_sheet(wb)  #類別二-間接蒸氣(汽電共生廠沒有做溫室氣體盤查)
    # sheet17.create_other_category2_sheet(wb)   #類別二-間接蒸氣(汽電共生廠有做溫室氣體盤查)
    sheet18.create_sheet18(wb)


    # 儲存 Excel 檔案
    output_filename = "碳盤查基準活動數據清冊.xlsx"
    wb.save(output_filename)
    print(f"Excel 檔案已生成: {output_filename}")

if __name__ == "__main__":
    main()
