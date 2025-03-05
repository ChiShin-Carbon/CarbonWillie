from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import style # 引入樣式模組

def create_non_indirect_steam_sheet(wb):
    """建立 '類別二-間接蒸氣(汽電共生廠沒有做溫室氣體盤查)' 工作表"""
    sheet_name = "類別二-間接蒸氣(汽電共生廠沒有做溫室氣體盤查)"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    title_cell = ws.cell(row=1, column=1, value="間接蒸氣(汽電共生廠沒有做溫室氣體盤查)")
    title_cell.alignment = style.center_alignment
    title_cell.fill = style.yellow_fill  # 標題背景色
    
    # 第二行：說明文字
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws.cell(row=2, column=1, value="").alignment = style.center_alignment
    
    # 第三行：表頭
    headers = ['原燃物料名稱', '使用量', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 灰色背景

    # 預設數據
    data = [
        {'fuel_name': '燃料煤', 'usage': 0, 'remark': ''},
        {'fuel_name': '燃料油', 'usage': 0, 'remark': ''},
        {'fuel_name': '天然氣(NG)', 'usage': 0, 'remark': ''},
        {'fuel_name': '液化天然氣(LNG)', 'usage': 0, 'remark': ''},
        {'fuel_name': '液化石油氣(LPG)', 'usage': 0, 'remark': ''},
        {'fuel_name': '液化石油氣(LPG)', 'usage': 0, 'remark': ''},
         {'fuel_name': '柴油', 'usage': 0, 'remark': ''},
        {'fuel_name': '高爐氣', 'usage': 0, 'remark': ''},
        {'fuel_name': '煉油氣', 'usage': 0, 'remark': ''},
        {'fuel_name': '石油焦', 'usage': 0, 'remark': ''},
        {'fuel_name': '全年蒸氣售出量', 'usage': 0, 'remark': ''},
        {'fuel_name': '公司本年度購買間接蒸氣量', 'usage': 0, 'remark': ''}
    ]

    row = 4  # 從第四行開始寫入資料
    for item in data:
        for col_idx, key in enumerate(['fuel_name', 'usage', 'remark'], start=1):
            ws.cell(row=row, column=col_idx, value=item.get(key, ''))
        row += 1

    # 第 6 行：欄位說明
    notes = ['', '數字', '文字(可不填)']
    for col_idx, note in enumerate(notes, start=1):
        ws.cell(row=row, column=col_idx, value=note)

    # 套用邊框
    for row_cells in ws.iter_rows(min_row=3, max_row=len(data)+3, min_col=1, max_col=len(headers)):
        for cell in row_cells:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2