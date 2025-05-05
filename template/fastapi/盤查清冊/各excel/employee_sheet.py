from openpyxl.utils import get_column_letter
import requests
from . import style  # 引入樣式模組

# API端點常數
API_BASE_URL = "http://localhost:8000"

def get_employee_data(year):
    """從API獲取指定年份的員工資料"""
    try:
        api_url = f"{API_BASE_URL}/employee_data_by_year/{year}"
        response = requests.get(api_url)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f"獲取員工資料失敗，狀態碼: {response.status_code}，錯誤訊息: {response.text}")
            return None
    except Exception as e:
        print(f"連接員工API時發生錯誤: {e}")
        return None

def create_work_hours_sheet(wb, data):
    """建立員工工作時數工作表"""
    sheet_name = "類別一-工作時數(員工)"
    ws = wb.create_sheet(title=sheet_name)
    
    # 設定標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    merged_cell = ws.cell(row=1, column=1, value="員工平均工作時數")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill  # 黃色背景
    
    # 第二行：標題列
    headers = ['月份', '員工數', '每日工時', '每月工作日數', '總加班時數', '總病假時數', '總事假時數', '總出差時數', '總婚喪時數', '總特休時數', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 灰色背景
    
    # 使用 API 資料填充
    month_map = {"01": "1月", "02": "2月", "03": "3月", "04": "4月", "05": "5月", "06": "6月",
                "07": "7月", "08": "8月", "09": "9月", "10": "10月", "11": "11月", "12": "12月"}
    
    # 設定起始行
    row_start = 3
    
    # 如果有資料，按月份排序並填入資料；如果沒有資料，則生成5個空行
    if data and len(data) > 0:
        # 按月份排序資料
        sorted_data = sorted(data, key=lambda x: x['month'])
        
        for i, item in enumerate(sorted_data, start=row_start):
            # 處理月份顯示格式
            month_display = month_map.get(item['month'], f"{item['month']}月")
            
            ws.cell(row=i, column=1, value=month_display)  # A 欄：月份
            ws.cell(row=i, column=2, value=item['employee_number'])  # B 欄：員工數
            ws.cell(row=i, column=3, value=item['daily_hours'])  # C 欄：每日工時
            ws.cell(row=i, column=4, value=item['workday'])  # D 欄：每月工作日數
            ws.cell(row=i, column=5, value=item.get('overtime', ''))  # E 欄：總加班時數
            ws.cell(row=i, column=6, value=item.get('sick_leave', ''))  # F 欄：總病假時數
            ws.cell(row=i, column=7, value=item.get('personal_leave', ''))  # G 欄：總事假時數
            ws.cell(row=i, column=8, value=item.get('business_trip', ''))  # H 欄：總出差時數
            ws.cell(row=i, column=9, value=item.get('wedding_and_funeral', ''))  # I 欄：總婚喪時數
            ws.cell(row=i, column=10, value=item.get('special_leave', ''))  # J 欄：總特休時數
            ws.cell(row=i, column=11, value=item.get('remark', ''))  # K 欄：備註
        
        row_end = row_start + len(sorted_data) - 1
    else:
        # 如果沒有資料或API連線失敗，生成5個空行
        for row_idx in range(row_start, row_start + 5):
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx, value="")
        
        row_end = row_start + 4  # 5個空行，末尾行編號為初始行+4
    
    # 設定總計行
    total_row = row_end + 1
    ws.cell(row=total_row, column=1, value="總計")
    ws.cell(row=total_row, column=2, value=f"=SUM(B{row_start}:B{row_end})")
    ws.cell(row=total_row, column=4, value=f"=SUM(D{row_start}:D{row_end})")
    ws.cell(row=total_row, column=5, value=f"=SUM(E{row_start}:E{row_end})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{row_start}:F{row_end})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G{row_start}:G{row_end})")
    ws.cell(row=total_row, column=8, value=f"=SUM(H{row_start}:H{row_end})")
    ws.cell(row=total_row, column=9, value=f"=SUM(I{row_start}:I{row_end})")
    ws.cell(row=total_row, column=10, value=f"=SUM(J{row_start}:J{row_end})")
    
    # 套用框線
    for row in ws.iter_rows(min_row=2, max_row=total_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border
    
    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2