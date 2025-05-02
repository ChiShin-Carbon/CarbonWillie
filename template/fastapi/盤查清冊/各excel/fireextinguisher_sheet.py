from openpyxl.utils import get_column_letter
import requests
import style  # 樣式模組

# API端點常數
API_BASE_URL = "http://localhost:8000"

def get_fireextinguisher_data():
    """從API獲取滅火器資料"""
    try:
        api_url = f"{API_BASE_URL}/fire_extinguisher_data"
        response = requests.get(api_url)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f"獲取滅火器資料失敗，狀態碼: {response.status_code}，錯誤訊息: {response.text}")
            return None
    except Exception as e:
        print(f"連接滅火器API時發生錯誤: {e}")
        return None

def create_fire_extinguisher_sheet(wb, data):
    """建立 '類別一-滅火器' 工作表"""
    sheet_name = "類別一-滅火器"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    merged_cell = ws.cell(row=1, column=1, value="滅火器")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill  # 標題背景色

    # 合併第二列說明文字
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=2, column=1, value="公司本年度二氧化碳滅火器之請購及換藥紀錄（無則免填）").alignment = style.center_alignment

    # 定義表頭
    headers = ['品名', '成分', '規格(重量)', '單位', '使用量(支)', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 標題背景色

    # 定義成分對應表 (將數字轉換為文字)
    ingredient_map = {
        0: "其他",
        1: "CO2",
        2: "HFC-227ea"
    }
    
    row_start = 4
    if data and len(data) > 0:
        # 填入資料
        for row_idx, item in enumerate(data, start=row_start):
            # API回傳欄位: item_name, ingredient, specification, remark
            ws.cell(row=row_idx, column=1, value=item['item_name'])  # 品名
            ws.cell(row=row_idx, column=2, value=ingredient_map.get(item['ingredient'], "其他"))  # 成分
            ws.cell(row=row_idx, column=3, value=item['specification'])  # 規格(重量)
            ws.cell(row=row_idx, column=4, value="公斤")  # 單位 (按要求統一使用公斤)
            ws.cell(row=row_idx, column=5, value="1")  # 使用量 (按要求統一使用1)
            ws.cell(row=row_idx, column=6, value=item.get('remark', ''))  # 備註
        row_end = row_start + len(data) - 1
    else:
        # 如果沒有資料或API連線失敗，生成5個空行
        for row_idx in range(row_start, row_start + 5):
            ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value="")
            ws.cell(row=row_idx, column=3, value="")
            ws.cell(row=row_idx, column=4, value="")
            ws.cell(row=row_idx, column=5, value="")
            ws.cell(row=row_idx, column=6, value="")
        row_end = row_start + 4  # 5個空行，末尾行編號為初始行+4
    
    # 欄位備註
    note_row = row_end + 1
    notes = ['文字', '選擇', '數字', '選擇', '數字', '文字']
    for col_idx, note in enumerate(notes, start=1):
        ws.cell(row=note_row, column=col_idx, value=note)

    # 套用黑色邊框
    for row in ws.iter_rows(min_row=3, max_row=row_end, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        adjusted_width = (max_length + 4) * 1.2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width