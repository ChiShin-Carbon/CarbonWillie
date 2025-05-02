from openpyxl.utils import get_column_letter
import requests
import style  # 樣式模組

# API端點常數
API_BASE_URL = "http://localhost:8000"

def get_vehicle_data(year):
    """從API獲取指定年份的車輛資料"""
    try:
        api_url = f"{API_BASE_URL}/vehicle_data_by_year/{year}"
        response = requests.get(api_url)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f"獲取車輛資料失敗，狀態碼: {response.status_code}，錯誤訊息: {response.text}")
            return None
    except Exception as e:
        print(f"連接車輛API時發生錯誤: {e}")
        return None

def create_sheet1(wb, data):
    """建立公務車(汽油)工作表"""
    sheet_name = "類別一-公務車(汽油)"
    ws = wb.create_sheet(title=sheet_name)

    # 設定標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    merged_cell = ws.cell(row=1, column=1, value="公務車(汽油)")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill

    # 第二行合併空白
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    # 標題列
    headers = ['發票日期', '油種', '單位', '公升數', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = style.gray_fill

    # 寫入資料或產生空行
    row_start = 4
    oil_map = {0: "車用汽油", 1: "車用柴油"}
    
    if data and len(data) > 0:
        # 如果有資料，則填入資料
        for i, item in enumerate(data, start=row_start):
            ws.cell(row=i, column=1, value=item['doc_date'])  # 發票日期
            ws.cell(row=i, column=2, value=oil_map.get(item['oil_species'], "未知"))  # 油種
            ws.cell(row=i, column=3, value="公升")  # 單位
            ws.cell(row=i, column=4, value=item['liters'])  # 公升數
            ws.cell(row=i, column=5, value=item.get('remark', ''))  # 備註
        row_end = row_start + len(data) - 1
    else:
        # 如果沒有資料或API連線失敗，生成5個空行
        for row_idx in range(row_start, row_start + 5):
            ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value="")
            ws.cell(row=row_idx, column=3, value="")
            ws.cell(row=row_idx, column=4, value="")
            ws.cell(row=row_idx, column=5, value="")
        row_end = row_start + 4  # 5個空行，末尾行編號為初始行+4

    # 套用框線
    for row in ws.iter_rows(min_row=3, max_row=row_end, min_col=1, max_col=5):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2