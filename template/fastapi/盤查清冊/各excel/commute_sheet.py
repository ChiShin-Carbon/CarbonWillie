from openpyxl.utils import get_column_letter
import requests
from . import style  # 引入樣式模組

# API端點常數
API_BASE_URL = "http://localhost:8000"

def get_commute_data(year):
    """從API獲取指定年份的員工通勤資料"""
    try:
        api_url = f"{API_BASE_URL}/commute_data_by_year/{year}"
        response = requests.get(api_url)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f"獲取通勤資料失敗，狀態碼: {response.status_code}，錯誤訊息: {response.text}")
            return None
    except Exception as e:
        print(f"連接通勤API時發生錯誤: {e}")
        return None

def create_commute_sheet(wb, data):
    """建立 '類別三-員工通勤' 工作表"""
    sheet_name = "類別三-員工通勤"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    merged_cell = ws.cell(row=1, column=1, value="員工通勤")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill  # 標題背景色

    # 定義表頭
    headers = ['交通方式', '公里數', '油種', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 表頭背景色

    # 交通方式對應表
    transportation_map = {
        1: "汽車",
        2: "機車",
        3: "公車",
        4: "捷運",
        5: "火車",
        6: "高鐵",
        7: "客運"
    }
    
    # 油種對應表
    oil_species_map = {
        1: "無",
        2: "汽油",
        3: "柴油"
    }
    
    # 如果有資料，則填入資料；如果沒有資料，則生成5個空行
    row_start = 3
    if data and len(data) > 0:
        for row_idx, item in enumerate(data, start=row_start):
            ws.cell(row=row_idx, column=1, value=transportation_map.get(item['transportation'], "未知"))
            ws.cell(row=row_idx, column=2, value=item['kilometers'])
            ws.cell(row=row_idx, column=3, value=oil_species_map.get(item['oil_species'], "未知"))
            ws.cell(row=row_idx, column=4, value=item.get('remark', ''))
        row_end = row_start + len(data) - 1
    else:
        # 如果沒有資料或API連線失敗，生成5個空行
        for row_idx in range(row_start, row_start + 5):
            ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value="")
            ws.cell(row=row_idx, column=3, value="")
            ws.cell(row=row_idx, column=4, value="")
        row_end = row_start + 4  # 5個空行，末尾行編號為初始行+4

    # 套用黑色邊框
    for row in ws.iter_rows(min_row=2, max_row=row_end, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2