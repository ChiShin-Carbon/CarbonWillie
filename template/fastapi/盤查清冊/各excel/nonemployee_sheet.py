from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import requests
from . import style  # 引入樣式模組

# 設定年份
# API端點常數
API_BASE_URL = "http://localhost:8000"

def get_nonemployee_data(year):
    try:
        api_url = f"{API_BASE_URL}/nonemployee_data_by_year/{year}"
        response = requests.get(api_url)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f"獲取非員工資料失敗，狀態碼: {response.status_code}，錯誤訊息: {response.text}")
            return None
    except Exception as e:
        print(f"連接非員工API時發生錯誤: {e}")
        return None

def create_nonemployee_work_hours_sheet(wb, data):
    """建立 '類別一-工作時數(非員工)' 工作表"""
    sheet_name = "類別一-工作時數(非員工)"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    merged_cell = ws.cell(row=1, column=1, value="工作時數(非員工)")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill  # 標題背景色

    # 定義表頭
    headers = ['月份', '人數', '總工時時數', '總工作人天', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 表頭背景色

    # 使用 API 資料填充
    month_map = {"01": "1月", "02": "2月", "03": "3月", "04": "4月", "05": "5月", "06": "6月",
                "07": "7月", "08": "8月", "09": "9月", "10": "10月", "11": "11月", "12": "12月"}
    
    row_start = 3
    if data and len(data) > 0:
        # 按月份排序資料
        sorted_data = sorted(data, key=lambda x: x['month'])
        
        for i, item in enumerate(sorted_data, start=row_start):
            # 處理月份顯示格式
            month_display = month_map.get(item['month'], f"{item['month']}月")
            
            ws.cell(row=i, column=1, value=month_display)  # A 欄：月份
            ws.cell(row=i, column=2, value=item['nonemployee_number'])  # B 欄：人數
            ws.cell(row=i, column=3, value=item['total_hours'])  # C 欄：總工時時數
            ws.cell(row=i, column=4, value=item['total_days'])  # D 欄：總工作人天
            ws.cell(row=i, column=5, value=item.get('remark', ''))  # E 欄：備註
        
        row_end = row_start + len(sorted_data) - 1
    else:
        # 如果沒有資料或API連線失敗，生成5個空行
        for row_idx in range(row_start, row_start + 5):
            ws.cell(row=row_idx, column=1, value="")  # 月份
            ws.cell(row=row_idx, column=2, value="")  # 人數
            ws.cell(row=row_idx, column=3, value="")  # 總工時時數
            ws.cell(row=row_idx, column=4, value="")  # 總工作人天
            ws.cell(row=row_idx, column=5, value="")  # 備註
        
        row_end = row_start + 4  # 5個空行，末尾行編號為初始行+4

    # 欄位備註
    note_row = row_end + 1
    notes = ['', '數字', '數字', '數字', '文字(可不填寫)']
    for col_idx, note in enumerate(notes, start=1):
        ws.cell(row=note_row, column=col_idx, value=note)

    # 套用黑色邊框
    for row in ws.iter_rows(min_row=2, max_row=row_end, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2