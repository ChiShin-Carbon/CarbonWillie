from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side
import style  # 引入樣式模組
import requests

# API端點常數
API_BASE_URL = "http://localhost:8000"

def get_generator_data(year):
    """從API獲取指定年份的緊急發電機資料"""
    try:
        # 確保year是有效的值且能轉換為整數
        if year is not None:
            # 處理各種可能的類型
            if isinstance(year, dict):
                # 如果是字典，嘗試取出年份相關的值
                # 常見的情況可能是 {'year': 2023} 或類似的格式
                # 嘗試幾種可能的鍵名
                for key in ['year', 'Year', 'value', 'id', 'date']:
                    if key in year and year[key] is not None:
                        year = year[key]
                        break
                # 如果沒有找到有效的鍵，使用字典的第一個值
                if isinstance(year, dict) and year:
                    year = next(iter(year.values()))
            elif isinstance(year, list) and len(year) > 0:
                # 如果是列表，取第一個元素
                year = year[0]
            
            # 最終轉換為整數
            if isinstance(year, (str, int, float)):
                year = int(year)
            else:
                # 如果轉換失敗，印出詳細資訊並返回None
                print(f"無法將 year 參數 ({type(year)}: {year}) 轉換為整數")
                return None
        
        # 構建API URL並發送請求
        api_url = f"{API_BASE_URL}/generator_data_by_year/{year}"
        response = requests.get(api_url)

        if response.status_code == 200:
            return response.json()
        elif response.status_code == 404:
            print(f"找不到{year}年的緊急發電機資料")
            return None
        else:
            print(f"獲取緊急發電機資料失敗，狀態碼: {response.status_code}，錯誤訊息: {response.text}")
            return None
            
    except Exception as e:
        print(f"連接緊急發電機API時發生錯誤: {e}")
        return None

def create_emergency_generator_sheet(wb, year=None):
    """建立 '類別一-緊急發電機' 工作表"""
    sheet_name = "類別一-緊急發電機"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = ws.cell(row=1, column=1, value="緊急發電機柴油填充紀錄")
    title_cell.alignment = style.center_alignment
    title_cell.fill = style.yellow_fill  # 標題背景色

    # 第二行：表頭
    headers = ['發票日期', '發票號碼', '使用量(公升)', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 灰色背景

    # 嘗試從API獲取數據
    data = None
    if year is not None:
        data = get_generator_data(year)
    
    # 填入數據
    row_start = 3  # 從第三行開始寫入資料
    if data and len(data) > 0:
        # 如果有資料，則填入資料
        for row_idx, item in enumerate(data, start=row_start):
            ws.cell(row=row_idx, column=1, value=item['doc_date'])  # 發票日期
            ws.cell(row=row_idx, column=2, value=item['doc_number'])  # 發票號碼
            ws.cell(row=row_idx, column=3, value=item['usage'])  # 使用量(公升)
            ws.cell(row=row_idx, column=4, value=item.get('remark', ''))  # 備註
        
        row_end = row_start + len(data) - 1
    else:
        # 如果沒有資料或API連線失敗，生成5個空行
        for row_idx in range(row_start, row_start + 5):
            ws.cell(row=row_idx, column=1, value="請輸入西元/月/日")
            ws.cell(row=row_idx, column=2, value="")
            ws.cell(row=row_idx, column=3, value=0)
            ws.cell(row=row_idx, column=4, value="none")
            
        row_end = row_start + 4  # 5個空行，末尾行編號為初始行+4

    # 第 row_end+1 行：欄位說明
    notes = ['日期', '', '數字', '文字']
    for col_idx, note in enumerate(notes, start=1):
        ws.cell(row=row_end+1, column=col_idx, value=note)

    # 套用邊框
    for row in ws.iter_rows(min_row=2, max_row=row_end, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2