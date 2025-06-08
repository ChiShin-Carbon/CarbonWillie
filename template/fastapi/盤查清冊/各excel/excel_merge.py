from fastapi import APIRouter, HTTPException, status, BackgroundTasks
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from connect.connect import connectDB
import os
import traceback
from openpyxl import Workbook
import requests
from typing import Optional
from datetime import datetime

# 取得當前檔案的目錄
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 設定 Excel 檔案保存路徑
UPLOAD_FOLDER = os.path.join(BASE_DIR, "..", "..", "..", "public", "original_excel")

# 確保資料夾存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# 建立 API 路由
excel_generator_router = APIRouter(tags=["Inventory API"])

# 回應模型
class GenerateExcelResponse(BaseModel):
    success: bool
    message: str
    file_path: Optional[str] = None
    year: Optional[int] = None

# 檢查文件狀態回應模型
class CheckExcelStatusResponse(BaseModel):
    exists: bool
    message: str
    file_path: Optional[str] = None

# 用於追蹤正在生成的Excel檔案
excel_generation_tasks = {}

# 獲取基準年份的函數
def get_baseline_year():
    """從資料庫獲取最新的基準年份"""
    conn = connectDB()
    if not conn:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="無法連接到資料庫")
    
    cursor = conn.cursor()
    try:
        # 查詢 baseline_id 最大的記錄
        cursor.execute("""
            SELECT TOP 1 baseline_id, cfv_start_date 
            FROM Baseline 
            ORDER BY baseline_id DESC
        """)
        
        result = cursor.fetchone()
        if not result:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="找不到任何 Baseline 記錄")
        
        baseline_id, cfv_start_date = result
        
        # 從日期字串中提取年份
        try:
            # 將 cfv_start_date 轉換為字符串以防它是 datetime 對象
            date_str = str(cfv_start_date)
            
            # 使用簡單的方法從日期字符串中提取年份
            # 假設格式為 YYYY-MM-DD 或者包含年份的其他格式
            if '-' in date_str:
                year = int(date_str.split('-')[0])
            elif '/' in date_str:
                year = int(date_str.split('/')[0])
            else:
                # 嘗試提取前4個字符作為年份
                year = int(date_str[:4])
                
            # 驗證提取的年份是否合理 (例如 1900-2100)
            if not (1900 <= year <= 2100):
                raise ValueError(f"提取的年份 {year} 不在合理範圍內")
            
            return year
                
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, 
                detail=f"無法從日期格式 '{cfv_start_date}' 中提取年份: {str(e)}"
            )
    
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"查詢發生錯誤: {e}")
    finally:
        conn.close()

# 根據 user_id 獲取 business_id 的函數
def get_business_id_by_user_id(user_id: int):
    """根據 user_id 從 users 資料表獲取 business_id"""
    conn = connectDB()
    if not conn:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="無法連接到資料庫")
    
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT business_id 
            FROM users 
            WHERE user_id = ?
        """, (user_id,))
        
        result = cursor.fetchone()
        if not result:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"找不到 user_id {user_id} 的使用者")
        
        return result[0]
    
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"查詢使用者資料時發生錯誤: {e}")
    finally:
        conn.close()

# 插入資料到 Inventory_Baseline 資料表
def insert_inventory_baseline(business_id: int, year: int, file_path: str):
    """插入資料到 Inventory_Baseline 資料表"""
    conn = connectDB()
    if not conn:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="無法連接到資料庫")
    
    cursor = conn.cursor()
    try:
        # 獲取當前時間
        created_at = datetime.now()
        
        cursor.execute("""
            INSERT INTO Inventory_Baseline (business_id, year, file_path, created_at)
            VALUES (?, ?, ?, ?)
        """, (business_id, year, file_path, created_at))
        
        conn.commit()
        print(f"成功插入 Inventory_Baseline 記錄: business_id={business_id}, year={year}, file_path={file_path}")
    
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"插入 Inventory_Baseline 資料時發生錯誤: {e}")
    finally:
        conn.close()

# 生成 Excel 檔案的函數
def generate_excel_file(year: int):
    """生成 Excel 檔案並保存到指定路徑"""
    try:
        wb = Workbook()
        wb.remove(wb.active)  # 移除預設的工作表

        # 匯入分工說明表模組
        from . import devision
        devision_data = devision.get_authorized_users_data(year)
        devision.create_sheetdevision(wb, devision_data)
        
        # 匯入車輛資料模組
        from . import vehicle_sheet
        vehicle_data = vehicle_sheet.get_vehicle_data(year)
        vehicle_sheet.create_sheet1(wb, vehicle_data)
        
        # 匯入員工資料模組
        from . import employee_sheet
        employee_data = employee_sheet.get_employee_data(year)
        employee_sheet.create_work_hours_sheet(wb, employee_data)
        
        # 匯入非員工資料模組
        from . import nonemployee_sheet
        nonemployee_data = nonemployee_sheet.get_nonemployee_data(year)
        nonemployee_sheet.create_nonemployee_work_hours_sheet(wb, nonemployee_data)

        # 匯入滅火器資料模組
        from . import fireextinguisher_sheet
        fireextinguisher_data = fireextinguisher_sheet.get_fireextinguisher_data()
        fireextinguisher_sheet.create_fire_extinguisher_sheet(wb, fireextinguisher_data)

        # 匯入廠內機具資料模組
        from . import machinery_sheet
        machinery_data = machinery_sheet.get_machinery_data(year)
        machinery_sheet.create_internal_machinery_sheet(wb, machinery_data)

        # 匯入緊急發電機資料模組
        from . import emergencygenerator_sheet
        generator_data = emergencygenerator_sheet.get_generator_data(year)
        emergencygenerator_sheet.create_emergency_generator_sheet(wb, generator_data)

        # 匯入電力使用量
        from . import electricityUsage_sheet
        electricity_data = electricityUsage_sheet.get_electricity_data(year)
        electricityUsage_sheet.create_electricity_usage_sheet(wb, electricity_data)

        # 匯入員工通勤資料模組
        from . import commute_sheet
        commute_data = commute_sheet.get_commute_data(year)
        commute_sheet.create_commute_sheet(wb, commute_data)

        # 匯入商務旅行資料模組
        from . import businesstrip_sheet
        businesstrip_data = businesstrip_sheet.get_businesstrip_data(year)
        businesstrip_sheet.create_businesstrip_sheet(wb, businesstrip_data)

        # 營運產生的廢棄物資料模組
        from . import operationalwaste_sheet
        operationalwaste_data = operationalwaste_sheet.get_operational_waste_data(year)
        operationalwaste_sheet.create_operational_waste_sheet(wb, operationalwaste_data)
        
        # 銷售產品的廢棄物資料模組
        from . import sellingwaste_sheet
        sellingwaste_data = sellingwaste_sheet.get_selling_waste_data(year)
        sellingwaste_sheet.create_sellingwaste_sheet(wb, sellingwaste_data)
        
        # 儲存 Excel 檔案
        output_filename = f"{year}盤查清冊.xlsx"
        output_path = os.path.join(UPLOAD_FOLDER, output_filename)
        wb.save(output_path)
        
        # 更新任務狀態為完成
        excel_generation_tasks[year] = {
            "status": "completed",
            "file_path": output_path
        }
        
        return output_path
    except Exception as e:
        # 更新任務狀態為失敗
        excel_generation_tasks[year] = {
            "status": "failed",
            "error": str(e)
        }
        error_msg = f"生成Excel檔案時發生錯誤: {e}\n{traceback.format_exc()}"
        raise Exception(error_msg)

# 非同步任務處理函數
def process_excel_generation(year: int, user_id: int):
    try:
        # 更新任務狀態為進行中
        excel_generation_tasks[year] = {
            "status": "processing",
            "file_path": None
        }
        
        # 生成 Excel 檔案
        output_path = generate_excel_file(year)
        
        # 獲取 business_id
        business_id = get_business_id_by_user_id(user_id)
        
        # 準備資料庫路徑格式
        db_file_path = f"/original_excel/{year}盤查清冊.xlsx"
        
        # 插入資料到 Inventory_Baseline 資料表
        insert_inventory_baseline(business_id, year, db_file_path)
        
        print(f"Excel 檔案已成功生成: {output_path}")
        print(f"資料已成功插入 Inventory_Baseline 資料表")
        
    except Exception as e:
        print(f"Excel 檔案生成或資料庫插入失敗: {e}")

# API 端點
@excel_generator_router.post("/generate_inventory_excel/{user_id}", response_model=GenerateExcelResponse)
async def generate_inventory_excel(user_id: int, background_tasks: BackgroundTasks):
    """
    生成碳盤查清冊 Excel 檔案，並保存到指定路徑，同時插入資料到 Inventory_Baseline 資料表
    """
    try:
        # 驗證 user_id 是否存在並獲取 business_id
        business_id = get_business_id_by_user_id(user_id)
        
        # 獲取基準年份
        year = get_baseline_year()
        if not year:
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                content={"success": False, "message": "無法獲取基準年份", "year": None, "file_path": None}
            )
        
        # 使用背景任務來生成 Excel 檔案並插入資料庫
        background_tasks.add_task(process_excel_generation, year, user_id)
        
        # 生成的檔案路徑
        output_filename = f"{year}盤查清冊.xlsx"
        output_path = os.path.join(UPLOAD_FOLDER, output_filename)
        
        return {
            "success": True,
            "message": f"已開始生成 {year} 年度碳盤查清冊，檔案生成後將保存至 {output_path}，並記錄至資料庫",
            "year": year,
            "file_path": output_path
        }
    
    except HTTPException as he:
        # 直接重新拋出 HTTP 例外
        raise he
    except Exception as e:
        error_detail = str(e)
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={
                "success": False,
                "message": f"生成 Excel 檔案時發生錯誤: {error_detail}",
                "year": None,
                "file_path": None
            }
        )

# 新增一個API端點用於檢查Excel檔案是否生成完成
@excel_generator_router.get("/check_excel_status/{year}", response_model=CheckExcelStatusResponse)
async def check_excel_status(year: int):
    """
    檢查指定年份的Excel檔案是否已生成完成
    """
    try:
        # 檢查任務狀態
        if year in excel_generation_tasks:
            task_info = excel_generation_tasks[year]
            
            if task_info["status"] == "completed":
                return {
                    "exists": True,
                    "message": f"{year}年度盤查清冊已成功生成",
                    "file_path": task_info["file_path"]
                }
            elif task_info["status"] == "failed":
                return {
                    "exists": False,
                    "message": f"{year}年度盤查清冊生成失敗: {task_info.get('error', '未知錯誤')}",
                    "file_path": None
                }
            else:  # processing
                return {
                    "exists": False,
                    "message": f"{year}年度盤查清冊正在生成中",
                    "file_path": None
                }
        
        # 如果沒有任務記錄，檢查文件是否存在
        output_filename = f"{year}盤查清冊.xlsx"
        output_path = os.path.join(UPLOAD_FOLDER, output_filename)
        
        if os.path.exists(output_path):
            # 文件存在但沒有任務記錄，可能是之前生成的
            return {
                "exists": True,
                "message": f"{year}年度盤查清冊已存在",
                "file_path": output_path
            }
        else:
            # 文件不存在且沒有任務記錄
            return {
                "exists": False,
                "message": f"{year}年度盤查清冊尚未生成",
                "file_path": None
            }
    
    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={
                "exists": False,
                "message": f"檢查Excel檔案狀態時發生錯誤: {str(e)}",
                "file_path": None
            }
        )