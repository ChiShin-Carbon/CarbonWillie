from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side
from . import style  # 引入樣式模組
import requests

# API端點常數
API_BASE_URL = "http://localhost:8000"

def get_electricity_data(year):
    """從API獲取指定年份的電力使用量資料"""
    try:
        # 確保year是整數
        year = int(year) if year is not None else None
        
        api_url = f"{API_BASE_URL}/electricity_data_by_year/{year}"
        response = requests.get(api_url)

        if response.status_code == 200:
            return response.json()
        elif response.status_code == 404:
            print(f"找不到{year}年的電力使用量資料")
            return None
        else:
            print(f"獲取電力使用量資料失敗，狀態碼: {response.status_code}，錯誤訊息: {response.text}")
            
    except Exception as e:
        print(f"連接電力使用量API時發生錯誤: {e}")
        return None

def create_electricity_usage_sheet(wb, data=None):
    """建立 '類別二-電力使用量' 工作表"""
    # 如果沒有提供資料，就使用空資料
    if data is None:
        data = []
    
    sheet_name = "類別二-電力使用量"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=1, value="電力使用量")
    title_cell.alignment = style.center_alignment
    title_cell.fill = style.yellow_fill

    # 第二行：說明文字
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=2, column=1, value="檢視往年資料").alignment = style.center_alignment

    # 第三行：表頭
    headers = ['收據月份', '用電期間(起)', '用電期間(迄)', '填寫類型', '當月總用電量或總金額', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = style.gray_fill

    # 填寫類型對應表
    electricity_type_map = {
        1: "總用電量",
        2: "總金額"
    }
    
    # 填入數據
    row_start = 4
    if data and len(data) > 0:
        for row_idx, item in enumerate(data, start=row_start):
            # 從日期中提取月份並加上"月"字
            doc_date = item['doc_date']
            month = doc_date.split('-')[1] if '-' in doc_date else ''
            month_display = f"{month}月" if month else ''
            
            ws.cell(row=row_idx, column=1, value=month_display)
            ws.cell(row=row_idx, column=2, value=item['period_start'])
            ws.cell(row=row_idx, column=3, value=item['period_end'])
            ws.cell(row=row_idx, column=4, value=electricity_type_map.get(item['electricity_type'], "未知"))
            ws.cell(row=row_idx, column=5, value=item['value'])
            ws.cell(row=row_idx, column=6, value='')  # 備註欄位預設為空
        
        row_end = row_start + len(data) - 1
        
        # 如果資料少於5筆，補充空行到5筆
        if len(data) < 5:
            for row_idx in range(row_start + len(data), row_start + 5):
                ws.cell(row=row_idx, column=1, value="選擇時期")
                ws.cell(row=row_idx, column=2, value="請輸入西元/月/日")
                ws.cell(row=row_idx, column=3, value="請輸入西元/月/日")
                ws.cell(row=row_idx, column=4, value="選擇填寫類型")
                ws.cell(row=row_idx, column=5, value=0)
                ws.cell(row=row_idx, column=6, value="")
            
            row_end = row_start + 4  # 確保至少有5行
    else:
        # 如果沒有資料或API連線失敗，生成5個空行
        for row_idx in range(row_start, row_start + 5):
            ws.cell(row=row_idx, column=1, value="選擇時期")
            ws.cell(row=row_idx, column=2, value="請輸入西元/月/日")
            ws.cell(row=row_idx, column=3, value="請輸入西元/月/日")
            ws.cell(row=row_idx, column=4, value="選擇填寫類型")
            ws.cell(row=row_idx, column=5, value=0)
            ws.cell(row=row_idx, column=6, value="")
            
        row_end = row_start + 4  # 5個空行，末尾行編號為初始行+4

    # 第 row_end+1 行：欄位說明
    notes = ['選擇', '日期', '日期', '選擇', '數字', '可不填']
    for col_idx, note in enumerate(notes, start=1):
        ws.cell(row=row_end+1, column=col_idx, value=note)

    # 套用邊框
    for row in ws.iter_rows(min_row=3, max_row=row_end, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2