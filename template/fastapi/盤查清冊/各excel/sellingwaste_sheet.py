from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import requests
from . import style  # 引入樣式模組

# API端點常數
API_BASE_URL = "http://localhost:8000"

def get_selling_waste_data(year):
    """從API獲取2024年的銷售產品廢棄物資料"""
    try:
        api_url = f"{API_BASE_URL}/selling_waste_data_by_year/{year}"
        response = requests.get(api_url)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f"獲取銷售產品廢棄物資料失敗，狀態碼: {response.status_code}，錯誤訊息: {response.text}")
            return None
    except Exception as e:
        print(f"連接銷售產品廢棄物API時發生錯誤: {e}")
        return None

def create_sellingwaste_sheet(wb, data=None):
    """建立 '類別三-銷售產品的廢棄物' 工作表
    
    參數:
    wb -- openpyxl Workbook物件
    data -- 銷售產品廢棄物資料列表，若為None則會使用API查詢資料
    """
    # 如果沒有提供資料，則從API獲取
    if data is None:
        data = []
    
    sheet_name = "類別三-銷售產品的廢棄物"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    merged_cell = ws.cell(row=1, column=1, value="銷售產品的廢棄物")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill  # 標題背景色

    # 定義表頭
    headers = ['廢棄物項目', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 表頭背景色

    # 如果有資料，則填入資料；如果沒有資料，則生成5個空行
    row_start = 3
    if data and len(data) > 0:
        for row_idx, item in enumerate(data, start=row_start):
            ws.cell(row=row_idx, column=1, value=item['waste_item'])
            ws.cell(row=row_idx, column=2, value=item.get('remark', ''))
        row_end = row_start + len(data) - 1
    else:
        # 如果沒有資料或API連線失敗，生成5個空行
        for row_idx in range(row_start, row_start + 5):
            ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value="")
        row_end = row_start + 4  # 5個空行，末尾行編號為初始行+4

    # 套用黑色邊框
    for row in ws.iter_rows(min_row=2, max_row=row_end, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2