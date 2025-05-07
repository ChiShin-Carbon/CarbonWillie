from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import requests
from . import style  # 引入樣式模組

# API端點常數
API_BASE_URL = "http://localhost:8000"

def get_authorized_users_data(year):
    """從API獲取指定年份的授權用戶資料"""
    try:
        api_url = f"{API_BASE_URL}/authorized_users_by_year/{year}"
        response = requests.get(api_url)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f"獲取授權用戶資料失敗，狀態碼: {response.status_code}，錯誤訊息: {response.text}")
            return None
    except Exception as e:
        print(f"連接授權用戶API時發生錯誤: {e}")
        return None

def create_sheetdevision(wb, data=None):
    """建立 '分工說明表' 工作表
    """
    # 如果沒有提供資料，則從API獲取
    if data is None:[]
    
    sheet_name = "分工說明表"
    ws = wb.create_sheet(title=sheet_name)

    # 定義標題
    headers = ['表單名稱', '部門', '姓名(主要填寫人)', '電子郵件', '電話及分機號碼', '平台帳號', '平台密碼', '平台權限', '備註']
    
    # 設定標題樣式
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = style.center_alignment
        cell.fill = style.yellow_fill

    # 各表單名稱
    sheet_names = [
        '類別一-公務車','類別一-滅火器',
        '類別一-工作時數(員工)', '類別一-工作時數(非員工)', '類別一-冷媒',
        '類別一-廠內機具','類別一-緊急發電機', 
        '類別二-電力使用量', 
        '類別三-員工通勤','類別三-商務旅行','類別三-營運產生廢棄物','類別三-銷售產品的廢棄物',
    ]
    
    # 創建一個表單名稱到完整名稱的映射
    table_name_to_full_name = {}
    for full_name in sheet_names:
        # 從完整名稱中提取表單名稱部分（類別後面的部分）
        parts = full_name.split('-')
        if len(parts) > 1:
            # 去掉文字中的括號和空格, 以便匹配
            simple_name = parts[1].strip().replace('(', '').replace(')', '')
            table_name_to_full_name[simple_name] = full_name
    
    # 處理API返回的資料
    if data and len(data) > 0:
        # 按照sheet_names的順序對資料進行分組
        organized_data = {}
        
        for sheet_entry in sheet_names:
            organized_data[sheet_entry] = []
        
        # 將每個API返回的數據分配到對應的表單項目
        for user_entry in data:
            # 從API返回的表單名稱中找到對應的完整表單名稱
            found = False
            for key, full_name in table_name_to_full_name.items():
                # 檢查API返回的table_name是否包含簡化名稱
                if key.lower() in user_entry['table_name'].lower():
                    organized_data[full_name].append(user_entry)
                    found = True
                    break
            
            # 如果沒找到對應的表單，尋找可能的匹配
            if not found:
                # 嘗試查找部分匹配
                for full_name in sheet_names:
                    if any(part.lower() in user_entry['table_name'].lower() for part in full_name.lower().split('-')):
                        organized_data[full_name].append(user_entry)
                        break
        
        # 填入資料
        current_row = 2
        
        for full_name in sheet_names:
            users = organized_data[full_name]
            
            if users:
                # 有資料的情況
                for user in users:
                    ws.cell(row=current_row, column=1, value=full_name)
                    ws.cell(row=current_row, column=2, value=user['department_name'])
                    ws.cell(row=current_row, column=3, value=user['username'])
                    ws.cell(row=current_row, column=4, value=user['email'])
                    ws.cell(row=current_row, column=5, value=user.get('telephone', ''))
                    ws.cell(row=current_row, column=6, value=user['address'])
                    ws.cell(row=current_row, column=7, value='')  # 密碼預設為空
                    ws.cell(row=current_row, column=8, value='')  # 權限預設為空
                    ws.cell(row=current_row, column=9, value='')  # 備註預設為空
                    current_row += 1
            else:
                # 沒有資料的情況，顯示空行
                ws.cell(row=current_row, column=1, value=full_name)
                for col in range(2, 10):
                    ws.cell(row=current_row, column=col, value='')
                current_row += 1
    else:
        # 如果API沒有返回任何資料，則為每個表單名稱創建一個空行
        current_row = 2
        for name in sheet_names:
            ws.cell(row=current_row, column=1, value=name)
            for col in range(2, 10):
                ws.cell(row=current_row, column=col, value='')
            current_row += 1
    
    

    # 最後：說明文字
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    merged_cell_explain = ws.cell(row=current_row, column=1, value="平台權限:\n"
            "1. 廠商負責人: 負責控管整個專案，可以查看及編輯所有填寫者的回覆。\n"
            "2. 主管階級: 檢視及編輯專案內的填寫資料。\n"
            "   *主管階級預設只能檢視專案內容，若需新增編輯權限，請在備註說明。*\n"
            "3. 一般使用者: 填寫自己所被分配的類別，如環安被分配到填寫冷媒，\n"
            "   那只能看到冷媒填寫頁面，不能查看其他類別 (如通勤、差旅) 之填寫頁面。"
    )
    merged_cell_explain.alignment = style.newline  # 自動換行
    merged_cell_explain.fill = style.yellow_fill  # 背景色

    # 套用黑色邊框 - 修改為使用實際的行數
    for row in ws.iter_rows(min_row=1, max_row=current_row-1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        adjusted_width = (max_length + 4) * 1.3  # 調整寬度
        if col_idx == 1:  # 第一欄加長
            adjusted_width = 50
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # 設定說明文字行的高度
    ws.row_dimensions[current_row].height = 120  # 設定較高的行高