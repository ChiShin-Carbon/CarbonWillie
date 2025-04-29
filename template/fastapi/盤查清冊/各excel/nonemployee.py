from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import requests
import style  # 引入樣式模組

# 設定年份
year = 2024
api_url = f"http://localhost:8000/nonemployee_data_by_year/{year}"  # 依實際 API 位址修改

def create_nonemployee_work_hours_sheet(wb, data):
    """建立 '類別一-工作時數(非員工)' 工作表"""
    sheet_name = "類別一-工作時數(非員工)"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    merged_cell = ws.cell(row=1, column=1, value="工作時數(非員工)")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill  # 標題背景色

    # 定義表頭
    headers = ['月份', '人數', '總工時時數', '總工作人天', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 表頭背景色

    # 使用 API 資料填充
    month_map = {"01": "1月", "02": "2月", "03": "3月", "04": "4月", "05": "5月", "06": "6月",
                "07": "7月", "08": "8月", "09": "9月", "10": "10月", "11": "11月", "12": "12月"}
    
    # 按月份排序資料
    sorted_data = sorted(data, key=lambda x: x['month'])
    
    for i, item in enumerate(sorted_data, start=3):
        # 處理月份顯示格式
        month_display = month_map.get(item['month'], f"{item['month']}月")
        
        ws.cell(row=i, column=1, value=month_display)  # A 欄：月份
        ws.cell(row=i, column=2, value=item['nonemployee_number'])  # B 欄：人數
        ws.cell(row=i, column=3, value=item['total_hours'])  # C 欄：總工時時數
        ws.cell(row=i, column=4, value=item['total_days'])  # D 欄：總工作人天
        ws.cell(row=i, column=5, value=item.get('remark', ''))  # E 欄：備註

    # 欄位備註
    note_row = len(sorted_data) + 3
    notes = ['', '數字', '數字', '數字', '文字(可不填寫)']
    for col_idx, note in enumerate(notes, start=1):
        ws.cell(row=note_row, column=col_idx, value=note)

    # 套用黑色邊框
    for row in ws.iter_rows(min_row=2, max_row=len(sorted_data)+2, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2


# 主程式區塊
def main():
    try:
        response = requests.get(api_url)
        
        if response.status_code == 200:
            data = response.json()
            
            wb = Workbook()
            wb.remove(wb.active)  # 移除預設工作表
            create_nonemployee_work_hours_sheet(wb, data)
            
            output_filename = f"nonemployee_hours_{year}.xlsx"
            wb.save(output_filename)
            print(f"Excel 檔案已生成: {output_filename}")
        else:
            print(f"API 請求失敗，狀態碼: {response.status_code}，錯誤訊息: {response.text}")
    except Exception as e:
        print(f"執行發生錯誤: {e}")

if __name__ == "__main__":
    main()