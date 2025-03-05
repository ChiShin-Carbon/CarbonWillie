from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 黑色框線（細線）
black_border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

# 置中對齊
center_alignment = Alignment(horizontal="center", vertical="center")

# 黃色背景（適用於標題列）
yellow_fill = PatternFill(fill_type="solid", fgColor="FFF4A460")

# 灰色背景（適用於表頭）
gray_fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD")

# 黃色背景2（適用於其他）
other_yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")

# 紅色字（適用於其他）
red_font = Font(color="FF0000")

# 自動換行(適用於sheetdevision分工說明表之文字說明)
newline = Alignment(wrap_text=True)