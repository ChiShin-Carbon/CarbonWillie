from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import requests
import style  # 樣式模組

# 設定年份
year = 2024
api_url = f"http://localhost:8000/vehicle_data_by_year/{year}"  # 依實際 API 位址修改

def create_sheet1(wb, data):
    sheet_name = "類別一-公務車(汽油)"
    ws = wb.create_sheet(title=sheet_name)

    # 設定標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    merged_cell = ws.cell(row=1, column=1, value="公務車(汽油)")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill

    # 第二行合併空白
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    # 標題列
    headers = ['發票日期', '油種', '單位', '公升數', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = style.gray_fill

    # 寫入 API 資料
    oil_map = {0: "車用汽油", 1: "車用柴油"}
    for i, item in enumerate(data, start=4):
        ws.cell(row=i, column=1, value=item['doc_date'])  # 發票日期
        ws.cell(row=i, column=2, value=oil_map.get(item['oil_species'], "未知"))  # 油種
        ws.cell(row=i, column=3, value="公升")  # 單位
        ws.cell(row=i, column=4, value=item['liters'])  # 公升數
        ws.cell(row=i, column=5, value=item.get('remark', ''))  # 備註

    # 套用框線
    for row in ws.iter_rows(min_row=3, max_row=len(data)+3, min_col=1, max_col=5):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2


# 主程式區塊
response = requests.get(api_url)

if response.status_code == 200:
    data = response.json()

    wb = Workbook()
    wb.remove(wb.active)  # 移除預設工作表
    create_sheet1(wb, data)

    output_filename = "sh1test_api.xlsx"
    wb.save(output_filename)
    print(f"Excel 檔案已生成: {output_filename}")
else:
    print(f"API 請求失敗，狀態碼: {response.status_code}，錯誤訊息: {response.text}")
