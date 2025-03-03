from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import style  # 引入樣式模組

def create_fixed_combustion_sheet(wb):
    """建立 '類別一-固定式燃燒' 工作表"""
    sheet_name = "類別一-固定式燃燒"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=1, value="公司固定式燃燒清冊")
    title_cell.alignment = style.center_alignment
    title_cell.fill = style.yellow_fill  # 標題背景色

    # 第二行：說明文字
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=2, column=1, value="檢視往年資料").alignment = style.center_alignment

    # 第三行：表頭
    headers = ['設備類型', '能源類型', '時期', '使用量', '單位', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 灰色背景

    # 填入數據
    data = [
        {'device_type': '', 'energy_type': '選擇能源類型', 'time': '選擇時期',
         'usage': 0, 'unit': '選擇單位', 'remark': 'none'}
    ] * 3  # 建立 3 行預設數據

    row = 4
    for item in data:
        for col_idx, key in enumerate(['device_type', 'energy_type', 'time', 'usage', 'unit', 'remark'], start=1):
            ws.cell(row=row, column=col_idx, value=item.get(key, ''))
        row += 1

    # 第 7 行：欄位說明
    notes = ['文字', '選擇', '', '數字', '選擇', '可不填']
    for col_idx, note in enumerate(notes, start=1):
        ws.cell(row=row, column=col_idx, value=note)

    # 套用邊框
    for row_cells in ws.iter_rows(min_row=3, max_row=len(data)+3, min_col=1, max_col=len(headers)):
        for cell in row_cells:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2
