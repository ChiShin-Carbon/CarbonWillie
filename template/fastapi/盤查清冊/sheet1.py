from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import style  # 引入樣式模組

def create_sheet1(wb):
    sheet_name = "類別一-公務車(汽油)"
    ws = wb.create_sheet(title=sheet_name)

    # 設定標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    merged_cell = ws.cell(row=1, column=1, value="公務車(汽油)")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill  # 黃色背景

    # 第二行
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    # 第三行：標題列
    headers = ['發票日期', '油種', '單位', '金額', '公升數', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 灰色背景

    # 填充月份與基本數據
    Doc_date = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月']
    for i, month in enumerate(Doc_date, start=4):
        ws.cell(row=i, column=1, value=month)  # A 欄：發票日期
        ws.cell(row=i, column=2, value='車用汽油')  # B 欄：油種

    # 套用框線
    for row in ws.iter_rows(min_row=3, max_row=len(Doc_date) + 3, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2
