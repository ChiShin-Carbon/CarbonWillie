from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import style  # 引入樣式模組

def create_nonemployee_work_hours_sheet(wb):
    """建立 '類別一-工作時數(非員工)' 工作表"""
    sheet_name = "類別一-工作時數(非員工)"
    ws = wb.create_sheet(title=sheet_name)

    # 合併標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    merged_cell = ws.cell(row=1, column=1, value="工作時數(非員工)")
    merged_cell.alignment = style.center_alignment
    merged_cell.fill = style.yellow_fill  # 標題背景色

    # 定義表頭
    headers = ['月份', '人數', '總工時時數', '總工作人天', '備註']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = style.gray_fill  # 表頭背景色

    # 填入月份
    months = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月']
    for row_idx, month in enumerate(months, start=3):
        ws.cell(row=row_idx, column=1, value=month)

    # 填入數據
    data = [
        {'nonemployee_number': 23, 'total_hours': 8, 'total_days': 21, 'remark': 'none'},
        {'nonemployee_number': 23, 'total_hours': 8, 'total_days': 15, 'remark': 'none'},
        {'nonemployee_number': 22, 'total_hours': 8, 'total_days': 23, 'remark': 'none'},
        {'nonemployee_number': 22, 'total_hours': 8, 'total_days': 19, 'remark': 'none'},
        {'nonemployee_number': 24, 'total_hours': 8, 'total_days': 21, 'remark': 'none'},
        {'nonemployee_number': 23, 'total_hours': 8, 'total_days': 21, 'remark': 'none'},
        {'nonemployee_number': 24, 'total_hours': 8, 'total_days': 20, 'remark': 'none'},
        {'nonemployee_number': 24, 'total_hours': 8, 'total_days': 23, 'remark': 'none'},
        {'nonemployee_number': 23, 'total_hours': 8, 'total_days': 21, 'remark': 'none'},
        {'nonemployee_number': 23, 'total_hours': 8, 'total_days': 20, 'remark': 'none'},
        {'nonemployee_number': 25, 'total_hours': 8, 'total_days': 22, 'remark': 'none'},
        {'nonemployee_number': 25, 'total_hours': 8, 'total_days': 22, 'remark': 'none'},
    ]

    for row_idx, entry in enumerate(data, start=3):
        for col_idx, key in enumerate(['nonemployee_number', 'total_hours', 'total_days', 'remark'], start=2):
            ws.cell(row=row_idx, column=col_idx, value=entry[key])

    # 欄位備註
    note_row = len(months) + 3
    notes = ['', '數字', '數字', '數字', '文字(可不填寫)']
    for col_idx, note in enumerate(notes, start=1):
        ws.cell(row=note_row, column=col_idx, value=note)

    # 套用黑色邊框
    for row in ws.iter_rows(min_row=2, max_row=len(data)+2, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = style.black_border

    # 自動調整欄寬
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = (max_length + 4) * 1.2
